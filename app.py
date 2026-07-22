import os
import sqlite3
import csv
import secrets
import smtplib
import json
import hashlib
import shutil
import tempfile
import zipfile
from email.message import EmailMessage
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, send_file, abort
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.exceptions import RequestEntityTooLarge
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DEFAULT_APP_ENV = 'development' if __name__ == '__main__' else 'production'
APP_ENV = os.environ.get('APP_ENV', DEFAULT_APP_ENV).strip().lower()
IS_PRODUCTION = APP_ENV == 'production'
DEMO_MODE = os.environ.get('APP_DEMO_MODE', '1' if DEFAULT_APP_ENV == 'development' else '0') == '1'

if os.environ.get('TRUST_PROXY', '0') == '1':
    # Enable only when the hosting provider supplies trusted proxy headers.
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)


def _env_int(name, default, minimum, maximum):
    try:
        value = int(os.environ.get(name, str(default)))
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(value, maximum))

# SECURITY SETTINGS
# The secret key protects sessions. It is generated once and stored locally,
# but on a live server you should set FLASK_SECRET_KEY in environment variables.
def _load_secret_key():
    env_key = os.environ.get('FLASK_SECRET_KEY')
    if env_key and len(env_key) >= 32:
        return env_key
    if IS_PRODUCTION:
        raise RuntimeError('FLASK_SECRET_KEY must be set to a random value of at least 32 characters in production.')
    key_file = os.path.join(BASE_DIR, '.secret_key')
    if os.path.exists(key_file):
        with open(key_file, 'r', encoding='utf-8') as f:
            return f.read().strip()
    key = secrets.token_urlsafe(48)
    with open(key_file, 'w', encoding='utf-8') as f:
        f.write(key)
    return key

app.secret_key = _load_secret_key()
MAX_UPLOAD_MB = _env_int('MAX_UPLOAD_MB', 50, 5, 100)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=IS_PRODUCTION,
    SESSION_COOKIE_NAME='kbss_session',
    PERMANENT_SESSION_LIFETIME=timedelta(minutes=45),
    MAX_CONTENT_LENGTH=MAX_UPLOAD_MB * 1024 * 1024,
    PREFERRED_URL_SCHEME='https' if IS_PRODUCTION else 'http',
)
DATA_DIR = os.path.abspath(os.environ.get('DATA_DIR', '').strip() or BASE_DIR)
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, 'school.db')
BACKUP_DIR = os.path.abspath(os.environ.get('BACKUP_DIR', '').strip() or os.path.join(DATA_DIR, 'backups'))
UPLOAD_ROOT = os.path.join(DATA_DIR, 'uploads')
MATERIAL_DIR = os.path.join(UPLOAD_ROOT, 'materials')
DOCUMENT_DIR = os.path.join(UPLOAD_ROOT, 'documents')
PROFILE_DIR = os.path.join(UPLOAD_ROOT, 'profile_pics')
GUIDANCE_VIDEO_DIR = os.path.join(UPLOAD_ROOT, 'guidance_videos')
BACKGROUND_DIR = os.path.join(UPLOAD_ROOT, 'backgrounds')
for d in [MATERIAL_DIR, DOCUMENT_DIR, PROFILE_DIR, GUIDANCE_VIDEO_DIR, BACKGROUND_DIR]:
    os.makedirs(d, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

DEPARTMENTS = {
    'mathematics': 'Mathematics Department',
    'natural_science': 'Natural Science Department',
    'social_science': 'Social Science Department',
    'computer_science': 'Computer Science Department',
    'business': 'Business Department',
    'home_economics': 'Home Economics Department',
    'language': 'Language Department'
}
ROLE_TO_DEPT = {
    'hod_mathematics': 'mathematics', 'hod_natural_science': 'natural_science',
    'hod_social_science': 'social_science', 'hod_computer_science': 'computer_science',
    'hod_business': 'business', 'hod_home_economics': 'home_economics', 'hod_language': 'language'
}
FULL_ACCESS_ROLES = ['headteacher', 'hr']
PORTAL_CONTROL_ROLES = ['headteacher', 'deputy_headteacher', 'hr']
GUIDANCE_ROLES = ['guidance_counselling']
STAFF_MANAGEMENT_FULL_ROLES = ['headteacher', 'deputy_headteacher', 'hr']
LEADERSHIP_ROLES = ['headteacher', 'deputy_headteacher', 'hr']
STAFF_ROLES = ['teacher', 'headteacher', 'deputy_headteacher', 'hr', 'guidance_counselling'] + list(ROLE_TO_DEPT.keys())
BACKGROUND_MANAGEMENT_ROLES = ['headteacher', 'deputy_headteacher'] + list(ROLE_TO_DEPT.keys())
CALENDAR_MANAGEMENT_ROLES = ['headteacher', 'deputy_headteacher', 'hr'] + list(ROLE_TO_DEPT.keys())
RECOVERY_ALLOWED_EMAIL = os.environ.get('RECOVERY_ALLOWED_EMAIL', '').strip().lower()
TERMS = ['Term 1', 'Term 2', 'Term 3']
GRADES = ['Form 1', 'Form 2', 'Form 3', 'Form 4', 'Form 5', 'Grade 10', 'Grade 11', 'Grade 12']

# Allowed file types protect the website from dangerous uploads such as .exe, .php, .bat.
ALLOWED_IMAGE_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
ALLOWED_DOCUMENT_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'txt', 'csv'}
ALLOWED_VIDEO_EXTENSIONS = {'mp4', 'webm', 'mov'}
LOGIN_MAX_ATTEMPTS = 5
LOGIN_LOCK_MINUTES = 15
RESULT_LOOKUP_MAX_ATTEMPTS = 8
RESULT_LOOKUP_LOCK_MINUTES = 15
MAX_GUIDANCE_VIDEO_MB = _env_int('MAX_GUIDANCE_VIDEO_MB', 50, 1, MAX_UPLOAD_MB)
MAX_GUIDANCE_VIDEO_BYTES = MAX_GUIDANCE_VIDEO_MB * 1024 * 1024
MAX_IMAGE_MB = _env_int('MAX_IMAGE_MB', 8, 1, MAX_UPLOAD_MB)
MAX_DOCUMENT_MB = _env_int('MAX_DOCUMENT_MB', 25, 1, MAX_UPLOAD_MB)
MAX_IMAGE_BYTES = MAX_IMAGE_MB * 1024 * 1024
MAX_DOCUMENT_BYTES = MAX_DOCUMENT_MB * 1024 * 1024
AUTO_BACKUP_ENABLED = os.environ.get('AUTO_BACKUP_ENABLED', '1' if IS_PRODUCTION else '0') == '1'
AUTO_BACKUP_HOURS = _env_int('AUTO_BACKUP_HOURS', 24, 1, 168)
BACKUP_RETENTION = _env_int('BACKUP_RETENTION', 30, 3, 365)


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def column_exists(cursor, table, column):
    cursor.execute(f'PRAGMA table_info({table})')
    return any(row[1] == column for row in cursor.fetchall())


DEMO_CREDENTIALS = {
    'head': 'head123', 'deputy': 'deputy123', 'hr': 'hr123',
    'teacher': 'teacher123', 'guidance': 'guidance123',
    'hod_math': 'math123', 'hod_natural': 'natural123',
    'hod_social': 'social123', 'hod_computer': 'computer123',
    'hod_business': 'business123', 'hod_home': 'home123',
    'hod_language': 'language123', 'student1': 'student123',
    'student2': 'student123',
}


def harden_production_accounts(cursor):
    """Disable known demo credentials and guarantee one secure bootstrap administrator."""
    if not IS_PRODUCTION or DEMO_MODE:
        return

    bootstrap_password = os.environ.get('INITIAL_ADMIN_PASSWORD', '')
    head = cursor.execute("SELECT * FROM users WHERE username='head'").fetchone()
    head_uses_default = bool(head and check_password_hash(head['password'], DEMO_CREDENTIALS['head']))

    if not head or head_uses_default:
        ok, message = password_is_strong(bootstrap_password)
        if not ok or len(bootstrap_password) < 12:
            raise RuntimeError(
                'Set INITIAL_ADMIN_PASSWORD to at least 12 characters with uppercase, lowercase and a number. '
                'It is required when installing the production upgrade or replacing the default Headteacher password.'
            )
        if head:
            cursor.execute(
                "UPDATE users SET password=?, must_change_password=1, is_active=1 WHERE id=?",
                (generate_password_hash(bootstrap_password), head['id'])
            )
        else:
            cursor.execute('''INSERT INTO users(
                username,password,role,full_name,position,email,must_change_password,is_active
            ) VALUES(?,?,?,?,?,?,1,1)''', (
                'head', generate_password_hash(bootstrap_password), 'headteacher',
                'Headteacher', 'Head Teacher', RECOVERY_ALLOWED_EMAIL
            ))

    for username, known_password in DEMO_CREDENTIALS.items():
        if username == 'head':
            continue
        user = cursor.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        if user and check_password_hash(user['password'], known_password):
            cursor.execute(
                'UPDATE users SET password=?, must_change_password=1, is_active=0 WHERE id=?',
                (generate_password_hash(secrets.token_urlsafe(32)), user['id'])
            )

    # Remove only the unmistakable demonstration pupil records shipped in older packages.
    sample_numbers = ('KBSS-001', 'KBSS-002', 'KBSS-003')
    placeholders = ','.join('?' for _ in sample_numbers)
    sample_rows = cursor.execute(
        f"SELECT id, student_number FROM students WHERE student_number IN ({placeholders}) AND full_name LIKE 'Sample Learner%'",
        sample_numbers
    ).fetchall()
    for sample in sample_rows:
        cursor.execute('DELETE FROM results WHERE student_id=?', (sample['id'],))
        cursor.execute('DELETE FROM result_download_logs WHERE student_id=? OR student_number=?', (sample['id'], sample['student_number']))
        cursor.execute("DELETE FROM users WHERE role='student' AND student_number=?", (sample['student_number'],))
        cursor.execute('DELETE FROM students WHERE id=?', (sample['id'],))
    cursor.execute("UPDATE password_reset_tokens SET used='Yes' WHERE used='No' AND length(token) != 64")


def init_db():
    conn = get_db(); c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL,
        full_name TEXT NOT NULL,
        position TEXT NOT NULL,
        department TEXT,
        phone TEXT,
        email TEXT,
        bio TEXT,
        address TEXT,
        qualification TEXT,
        profile_picture TEXT,
        student_number TEXT,
        must_change_password INTEGER NOT NULL DEFAULT 0,
        is_active INTEGER NOT NULL DEFAULT 1,
        profile_updated_at TEXT
    )''')
    # Safe upgrades for older database versions
    for col, coltype in [
        ('address','TEXT'), ('qualification','TEXT'), ('profile_picture','TEXT'),
        ('student_number','TEXT'), ('must_change_password','INTEGER NOT NULL DEFAULT 0'),
        ('is_active','INTEGER NOT NULL DEFAULT 1'), ('profile_updated_at','TEXT')
    ]:
        if not column_exists(c, 'users', col):
            c.execute(f'ALTER TABLE users ADD COLUMN {col} {coltype}')

    c.execute('''CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_number TEXT UNIQUE NOT NULL,
        full_name TEXT NOT NULL,
        grade TEXT NOT NULL,
        class_name TEXT NOT NULL,
        gender TEXT,
        parent_phone TEXT,
        class_teacher TEXT,
        created_at TEXT NOT NULL
    )''')
    if not column_exists(c, 'students', 'class_teacher'):
        c.execute('ALTER TABLE students ADD COLUMN class_teacher TEXT')

    c.execute('''CREATE TABLE IF NOT EXISTS portal_settings (
        setting_key TEXT PRIMARY KEY,
        setting_value TEXT NOT NULL,
        updated_by TEXT,
        updated_at TEXT
    )''')
    if not c.execute('SELECT setting_value FROM portal_settings WHERE setting_key=?', ('student_results_active',)).fetchone():
        c.execute('INSERT INTO portal_settings(setting_key, setting_value, updated_by, updated_at) VALUES(?,?,?,?)',
                  ('student_results_active', 'active' if DEMO_MODE else 'inactive', 'System', datetime.now().strftime('%Y-%m-%d %H:%M')))

    c.execute('''CREATE TABLE IF NOT EXISTS results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        subject TEXT NOT NULL,
        grade TEXT NOT NULL,
        term TEXT NOT NULL,
        academic_year TEXT NOT NULL,
        test1 REAL DEFAULT 0,
        test2 REAL DEFAULT 0,
        end_term REAL DEFAULT 0,
        total REAL DEFAULT 0,
        average REAL DEFAULT 0,
        comment TEXT,
        entered_by TEXT NOT NULL,
        entered_at TEXT NOT NULL,
        FOREIGN KEY(student_id) REFERENCES students(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS result_download_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        student_number TEXT,
        student_name TEXT,
        grade TEXT,
        class_name TEXT,
        downloaded_by TEXT,
        downloader_role TEXT,
        downloaded_at TEXT NOT NULL
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS materials (id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL, description TEXT, department TEXT NOT NULL, filename TEXT NOT NULL, uploaded_by TEXT NOT NULL, uploaded_at TEXT NOT NULL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS documents (id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL, description TEXT, category TEXT NOT NULL, filename TEXT NOT NULL, uploaded_by TEXT NOT NULL, uploaded_at TEXT NOT NULL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS donor_needs (id INTEGER PRIMARY KEY AUTOINCREMENT, item TEXT NOT NULL, description TEXT, quantity TEXT, priority TEXT NOT NULL, estimated_cost TEXT, status TEXT NOT NULL DEFAULT 'Needed', created_at TEXT NOT NULL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS donation_pledges (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        need_id INTEGER,
        donor_name TEXT NOT NULL,
        email TEXT,
        phone TEXT,
        donation_type TEXT NOT NULL,
        amount TEXT,
        item_description TEXT,
        payment_method TEXT,
        message TEXT,
        status TEXT NOT NULL DEFAULT 'Pending',
        pledged_at TEXT NOT NULL,
        FOREIGN KEY(need_id) REFERENCES donor_needs(id)
    )''')
    c.execute("""CREATE TABLE IF NOT EXISTS guidance_posts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        category TEXT NOT NULL,
        message TEXT NOT NULL,
        video_filename TEXT,
        posted_by TEXT NOT NULL,
        posted_at TEXT NOT NULL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS website_backgrounds (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        filename TEXT NOT NULL,
        uploaded_by TEXT NOT NULL,
        uploaded_role TEXT NOT NULL,
        is_active TEXT NOT NULL DEFAULT 'Yes',
        uploaded_at TEXT NOT NULL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS website_content (
        content_key TEXT PRIMARY KEY,
        content_value TEXT NOT NULL,
        updated_by TEXT,
        updated_at TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS school_calendar (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        event_date TEXT NOT NULL,
        start_time TEXT,
        end_time TEXT,
        category TEXT NOT NULL,
        venue TEXT,
        audience TEXT,
        description TEXT,
        created_by TEXT NOT NULL,
        created_at TEXT NOT NULL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS staff_returns (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER UNIQUE NOT NULL,
        full_name TEXT NOT NULL,
        employee_number TEXT,
        nrc_number TEXT,
        gender TEXT,
        date_of_birth TEXT,
        phone TEXT,
        email TEXT,
        address TEXT,
        position TEXT,
        department TEXT,
        subjects_taught TEXT,
        classes_taught TEXT,
        highest_qualification TEXT,
        professional_qualification TEXT,
        years_experience TEXT,
        date_first_appointed TEXT,
        date_joined_school TEXT,
        employment_status TEXT,
        tsc_number TEXT,
        next_of_kin TEXT,
        next_of_kin_phone TEXT,
        remarks TEXT,
        submitted_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(user_id) REFERENCES users(id)
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS staff_return_fields (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        field_label TEXT NOT NULL,
        section TEXT NOT NULL DEFAULT 'Additional Details Required by HR',
        field_type TEXT NOT NULL DEFAULT 'text',
        options TEXT,
        is_required TEXT NOT NULL DEFAULT 'No',
        is_active TEXT NOT NULL DEFAULT 'Yes',
        display_order INTEGER NOT NULL DEFAULT 1,
        created_by TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS staff_return_answers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        staff_return_id INTEGER NOT NULL,
        field_id INTEGER NOT NULL,
        answer_value TEXT,
        updated_at TEXT NOT NULL,
        UNIQUE(staff_return_id, field_id),
        FOREIGN KEY(staff_return_id) REFERENCES staff_returns(id),
        FOREIGN KEY(field_id) REFERENCES staff_return_fields(id)
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS staff_return_core_fields (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        field_key TEXT UNIQUE NOT NULL,
        field_label TEXT NOT NULL,
        section TEXT NOT NULL,
        is_required TEXT NOT NULL DEFAULT 'No',
        is_active TEXT NOT NULL DEFAULT 'Yes',
        display_order INTEGER NOT NULL DEFAULT 1,
        updated_at TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS password_audit_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        target_user_id INTEGER,
        target_username TEXT,
        target_full_name TEXT,
        target_role TEXT,
        action TEXT NOT NULL,
        changed_by_user_id INTEGER,
        changed_by_name TEXT,
        changed_by_role TEXT,
        changed_at TEXT NOT NULL,
        note TEXT,
        FOREIGN KEY(target_user_id) REFERENCES users(id),
        FOREIGN KEY(changed_by_user_id) REFERENCES users(id)
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS password_reset_tokens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        email TEXT NOT NULL,
        token TEXT UNIQUE NOT NULL,
        role TEXT NOT NULL,
        expires_at TEXT NOT NULL,
        used TEXT NOT NULL DEFAULT 'No',
        created_at TEXT NOT NULL,
        used_at TEXT,
        FOREIGN KEY(user_id) REFERENCES users(id)
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS password_reset_email_outbox (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        username TEXT NOT NULL,
        email TEXT NOT NULL,
        reset_link TEXT NOT NULL,
        status TEXT NOT NULL,
        created_at TEXT NOT NULL,
        sent_at TEXT,
        details TEXT,
        FOREIGN KEY(user_id) REFERENCES users(id)
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS security_audit_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_type TEXT NOT NULL,
        username TEXT,
        role TEXT,
        ip_address TEXT,
        user_agent TEXT,
        details TEXT,
        created_at TEXT NOT NULL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS login_attempts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        ip_address TEXT,
        success TEXT NOT NULL,
        attempted_at TEXT NOT NULL
    )""")
    default_core_fields = [
        ('full_name','Full Name','Personal Details','Yes','Yes',1),
        ('employee_number','Employee / Staff Number','Personal Details','No','Yes',2),
        ('nrc_number','NRC Number','Personal Details','No','Yes',3),
        ('gender','Gender','Personal Details','No','Yes',4),
        ('date_of_birth','Date of Birth','Personal Details','No','Yes',5),
        ('phone','Phone Number','Personal Details','No','Yes',6),
        ('email','Email Address','Personal Details','No','Yes',7),
        ('address','Residential Address','Personal Details','No','Yes',8),
        ('position','Position','Employment Details','No','Yes',9),
        ('department','Department','Employment Details','No','Yes',10),
        ('subjects_taught','Subjects Taught','Employment Details','No','Yes',11),
        ('classes_taught','Classes Taught','Employment Details','No','Yes',12),
        ('highest_qualification','Highest Academic Qualification','Employment Details','No','Yes',13),
        ('professional_qualification','Professional Qualification','Employment Details','No','Yes',14),
        ('years_experience','Years of Experience','Employment Details','No','Yes',15),
        ('employment_status','Employment Status','Employment Details','No','Yes',16),
        ('date_first_appointed','Date First Appointed','Employment Details','No','Yes',17),
        ('date_joined_school','Date Joined This School','Employment Details','No','Yes',18),
        ('tsc_number','TSC / Professional Number','Employment Details','No','Yes',19),
        ('next_of_kin','Next of Kin Name','Next of Kin','No','Yes',20),
        ('next_of_kin_phone','Next of Kin Phone','Next of Kin','No','Yes',21),
        ('remarks','Remarks','Other Details','No','Yes',22),
    ]
    for item in default_core_fields:
        c.execute("""INSERT OR IGNORE INTO staff_return_core_fields(field_key, field_label, section, is_required, is_active, display_order, updated_at)
                     VALUES(?,?,?,?,?,?,?)""", (*item, datetime.now().strftime('%Y-%m-%d %H:%M')))

    default_content = {
        'about_intro': 'Kafubu Block Secondary School is committed to providing quality education, discipline, innovation and learner-centred teaching. This website supports communication between school management, departments, teachers, learners, parents and donors.',
        'vision': 'To be a centre of academic excellence and responsible citizenship.',
        'mission': 'To provide quality secondary education through teamwork, discipline and effective use of teaching and learning materials.',
        'core_values': 'Integrity, hard work, accountability, respect, creativity and service.',
        'motto': 'Knowledge, Discipline and Service',
        'contact_phone': '+260 XXX XXX XXX',
        'contact_email': 'info@kafubublock.edu.zm',
        'office_hours': 'Monday to Friday, 07:30 - 16:30',
        'school_location': 'Kafubu Block Secondary School, Zambia',
        'map_query': 'Kafubu Block Secondary School, Zambia',
        'contact_message': 'Use the details below for enquiries, donations, school calendar information and general school communication.'
    }
    for key, value in default_content.items():
        if not c.execute('SELECT content_key FROM website_content WHERE content_key=?', (key,)).fetchone():
            c.execute('INSERT INTO website_content(content_key, content_value, updated_by, updated_at) VALUES(?,?,?,?)',
                      (key, value, 'System', datetime.now().strftime('%Y-%m-%d %H:%M')))

    users = [
        ('head', 'head123', 'headteacher', 'Mr/Ms Head Teacher', 'Head Teacher', None, 'Official school leader responsible for policy, administration and academic standards.'),
        ('deputy', 'deputy123', 'deputy_headteacher', 'Mr/Ms Deputy Head Teacher', 'Deputy Head Teacher', None, 'Supports school operations, discipline, teacher supervision and learner welfare.'),
        ('hr', 'hr123', 'hr', 'Human Resource Officer', 'Human Resource Officer', None, 'Manages staff records, welfare, leave documents and HR-related communication.'),
        ('teacher', 'teacher123', 'teacher', 'Subject Teacher', 'Subject Teacher', None, 'Subject teacher account for downloading departmental materials, entering learner results and accessing official documents.'),
        ('guidance', 'guidance123', 'guidance_counselling', 'Guidance and Counselling Teacher', 'Guidance and Counselling Teacher', None, 'Provides counselling support and posts educational guidance videos and messages for learners.'),
        ('hod_math', 'math123', 'hod_mathematics', 'HOD Mathematics', 'Head of Department', 'mathematics', 'Coordinates Mathematics teaching, assessment, records and learner support.'),
        ('hod_natural', 'natural123', 'hod_natural_science', 'HOD Natural Science', 'Head of Department', 'natural_science', 'Coordinates Biology, Chemistry and Physics teaching materials.'),
        ('hod_social', 'social123', 'hod_social_science', 'HOD Social Science', 'Head of Department', 'social_science', 'Coordinates Civic Education, History, Geography and Religious Education materials.'),
        ('hod_computer', 'computer123', 'hod_computer_science', 'HOD Computer Science', 'Head of Department', 'computer_science', 'Coordinates Computer Studies, ICT resources and digital learning.'),
        ('hod_business', 'business123', 'hod_business', 'HOD Business', 'Head of Department', 'business', 'Coordinates Commerce, Accounts, Business Studies and entrepreneurship materials.'),
        ('hod_home', 'home123', 'hod_home_economics', 'HOD Home Economics', 'Head of Department', 'home_economics', 'Coordinates Food and Nutrition, Design and Technology, and practical work.'),
        ('hod_language', 'language123', 'hod_language', 'HOD Language', 'Head of Department', 'language', 'Coordinates English, local languages, literacy and communication skills.'),
        ('student1', 'student123', 'student', 'Sample Learner One', 'Pupil', None, 'Pupil portal account for viewing Test 1, Test 2 and End Term results.'),
        ('student2', 'student123', 'student', 'Sample Learner Two', 'Pupil', None, 'Pupil portal account for viewing Test 1, Test 2 and End Term results.'),
    ] if DEMO_MODE else []
    for u in users:
        if not c.execute('SELECT id FROM users WHERE username=?', (u[0],)).fetchone():
            c.execute('''INSERT INTO users(username,password,role,full_name,position,department,bio,email,phone,qualification,address,student_number)
                         VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',
                      (u[0], generate_password_hash(u[1]), u[2], u[3], u[4], u[5], u[6], f'{u[0]}@kafubublock.edu.zm', '+260 XXX XXX XXX', '', '', 'KBSS-001' if u[0]=='student1' else ('KBSS-002' if u[0]=='student2' else None)))
    if DEMO_MODE:
        c.execute("UPDATE users SET full_name=?, position=?, bio=? WHERE username=?",
                  ('Subject Teacher', 'Subject Teacher', 'Subject teacher account for downloading departmental materials, entering learner results and accessing official documents.', 'teacher'))
        c.execute('UPDATE users SET student_number=? WHERE username=? AND (student_number IS NULL OR student_number="")', ('KBSS-001', 'student1'))
        c.execute('UPDATE users SET student_number=? WHERE username=? AND (student_number IS NULL OR student_number="")', ('KBSS-002', 'student2'))
    if RECOVERY_ALLOWED_EMAIL:
        c.execute('UPDATE users SET email=? WHERE username IN (?,?)', (RECOVERY_ALLOWED_EMAIL, 'head', 'hr'))

    sample_students = [
        ('KBSS-001', 'Sample Learner One', 'Form 1', '1A', 'Female', '+260 XXX XXX XXX', 'Mr/Ms Class Teacher 1A'),
        ('KBSS-002', 'Sample Learner Two', 'Form 2', '2A', 'Male', '+260 XXX XXX XXX', 'Mr/Ms Class Teacher 2A'),
        ('KBSS-003', 'Sample Learner Three', 'Form 3', '3A', 'Female', '+260 XXX XXX XXX', 'Mr/Ms Class Teacher 3A'),
    ] if DEMO_MODE else []
    for st in sample_students:
        if not c.execute('SELECT id FROM students WHERE student_number=?', (st[0],)).fetchone():
            c.execute('INSERT INTO students(student_number,full_name,grade,class_name,gender,parent_phone,class_teacher,created_at) VALUES(?,?,?,?,?,?,?,?)', (*st, datetime.now().strftime('%Y-%m-%d %H:%M')))
        else:
            c.execute('UPDATE students SET class_teacher=COALESCE(NULLIF(class_teacher,""), ?) WHERE student_number=?', (st[6], st[0]))

    if DEMO_MODE and not c.execute('SELECT id FROM donor_needs LIMIT 1').fetchone():
        needs = [
            ('Science laboratory equipment','Beakers, test tubes, burners and safety goggles for learner practicals.','Assorted','High','K35,000'),
            ('Computer laboratory computers','Desktop computers or laptops for ICT and Computer Science lessons.','25 computers','High','K180,000'),
            ('Library books','Updated textbooks, novels, dictionaries and revision books.','300 books','Medium','K45,000'),
            ('Sports equipment','Footballs, netballs, jerseys, cones and first aid kits.','Assorted','Medium','K15,000'),
            ('Classroom furniture','Desks and chairs to improve the learning environment.','100 desks','High','K80,000'),
        ]
        for n in needs:
            c.execute('INSERT INTO donor_needs(item,description,quantity,priority,estimated_cost,created_at) VALUES(?,?,?,?,?,?)', (*n, datetime.now().strftime('%Y-%m-%d %H:%M')))
    if DEMO_MODE and not c.execute('SELECT id FROM guidance_posts LIMIT 1').fetchone():
        c.execute("""INSERT INTO guidance_posts(title,category,message,video_filename,posted_by,posted_at) VALUES(?,?,?,?,?,?)""",
                  ('Welcome to Guidance and Counselling', 'Learner Support', 'This section will be used to share advice, study skills, discipline guidance, career information and counselling messages for learners.', None, 'Guidance and Counselling Teacher', datetime.now().strftime('%Y-%m-%d %H:%M')))
    if DEMO_MODE and not c.execute('SELECT id FROM school_calendar LIMIT 1').fetchone():
        calendar_events = [
            ('Opening Day', '2026-01-12', '07:30', '12:30', 'Term Event', 'School Campus', 'All learners and staff', 'Official opening of the school term and orientation for learners.'),
            ('Mid-Term Tests', '2026-02-16', '08:00', '15:30', 'Assessment', 'Classrooms', 'All learners', 'Beginning of mid-term tests. Learners must come prepared with required materials.'),
            ('Parents Meeting', '2026-03-06', '09:00', '12:00', 'Meeting', 'School Hall', 'Parents and guardians', 'Meeting to discuss learner performance, discipline and school development.'),
            ('End of Term Examinations', '2026-03-23', '08:00', '15:30', 'Assessment', 'Classrooms', 'All learners', 'Start of end-of-term examinations.'),
        ]
        for ev in calendar_events:
            c.execute("""INSERT INTO school_calendar(title,event_date,start_time,end_time,category,venue,audience,description,created_by,created_at)
                         VALUES(?,?,?,?,?,?,?,?,?,?)""", (*ev, 'System', datetime.now().strftime('%Y-%m-%d %H:%M')))
    harden_production_accounts(c)
    conn.commit(); conn.close()


@app.context_processor
def inject_globals():
    return dict(DEPARTMENTS=DEPARTMENTS, TERMS=TERMS, GRADES=GRADES, FULL_ACCESS_ROLES=FULL_ACCESS_ROLES, STAFF_MANAGEMENT_FULL_ROLES=STAFF_MANAGEMENT_FULL_ROLES, LEADERSHIP_ROLES=LEADERSHIP_ROLES, STAFF_ROLES=STAFF_ROLES, GUIDANCE_ROLES=GUIDANCE_ROLES, PORTAL_CONTROL_ROLES=PORTAL_CONTROL_ROLES, session=session, student_results_active=is_student_results_active(), active_backgrounds=get_active_backgrounds(), BACKGROUND_MANAGEMENT_ROLES=BACKGROUND_MANAGEMENT_ROLES, CALENDAR_MANAGEMENT_ROLES=CALENDAR_MANAGEMENT_ROLES, website_content=get_website_content(), MAX_GUIDANCE_VIDEO_MB=MAX_GUIDANCE_VIDEO_MB)



def get_setting(key, default=''):
    conn = get_db()
    row = conn.execute('SELECT setting_value FROM portal_settings WHERE setting_key=?', (key,)).fetchone()
    conn.close()
    return row['setting_value'] if row else default


def is_student_results_active():
    try:
        return get_setting('student_results_active', 'active') == 'active'
    except sqlite3.OperationalError:
        return True


def get_active_backgrounds():
    """Return active website background images for the public site slideshow."""
    try:
        conn = get_db()
        rows = conn.execute("SELECT * FROM website_backgrounds WHERE is_active='Yes' ORDER BY uploaded_at DESC").fetchall()
        conn.close()
        return rows
    except sqlite3.OperationalError:
        return []


def get_website_content():
    """Load editable public website content such as vision, mission, values and motto."""
    defaults = {
        'about_intro': 'Kafubu Block Secondary School is committed to providing quality education, discipline, innovation and learner-centred teaching. This website supports communication between school management, departments, teachers, learners, parents and donors.',
        'vision': 'To be a centre of academic excellence and responsible citizenship.',
        'mission': 'To provide quality secondary education through teamwork, discipline and effective use of teaching and learning materials.',
        'core_values': 'Integrity, hard work, accountability, respect, creativity and service.',
        'motto': 'Knowledge, Discipline and Service',
        'contact_phone': '+260 XXX XXX XXX',
        'contact_email': 'info@kafubublock.edu.zm',
        'office_hours': 'Monday to Friday, 07:30 - 16:30',
        'school_location': 'Kafubu Block Secondary School, Zambia',
        'map_query': 'Kafubu Block Secondary School, Zambia',
        'contact_message': 'Use the details below for enquiries, donations, school calendar information and general school communication.'
    }
    try:
        conn = get_db()
        rows = conn.execute('SELECT content_key, content_value FROM website_content').fetchall()
        conn.close()
        content = defaults.copy()
        for row in rows:
            content[row['content_key']] = row['content_value']
        return content
    except sqlite3.OperationalError:
        return defaults


def get_client_ip():
    # ProxyFix supplies the correct address only when TRUST_PROXY=1.
    return request.remote_addr or ''


def log_security_event(event_type, username=None, role=None, details=''):
    try:
        conn = get_db()
        conn.execute('''INSERT INTO security_audit_logs(event_type, username, role, ip_address, user_agent, details, created_at)
                        VALUES(?,?,?,?,?,?,?)''',
                     (event_type, username or session.get('username'), role or session.get('role'), get_client_ip(),
                      request.headers.get('User-Agent','')[:250], details[:1000], datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit(); conn.close()
    except Exception:
        pass


def is_allowed_file(filename, allowed_extensions):
    return bool(filename and '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions)


def unique_safe_filename(original_name):
    safe = secure_filename(original_name)
    if not safe:
        safe = 'uploaded_file'
    return datetime.now().strftime('%Y%m%d%H%M%S_%f_') + safe


def uploaded_file_within_limit(file, maximum_bytes):
    current_position = file.stream.tell()
    file.stream.seek(0, os.SEEK_END)
    size = file.stream.tell()
    file.stream.seek(current_position)
    return size <= maximum_bytes


def generate_temporary_password():
    # Easy to type, while still containing uppercase, lowercase and numbers.
    return f"Kb{secrets.randbelow(900000) + 100000}{secrets.choice('ABCDEFGHJKLMNPQRSTUVWXYZ')}{secrets.choice('abcdefghijkmnopqrstuvwxyz')}"


def password_is_strong(password):
    if len(password) < 8:
        return False, 'Password must have at least 8 characters.'
    if not any(ch.isupper() for ch in password):
        return False, 'Password must include at least one capital letter.'
    if not any(ch.islower() for ch in password):
        return False, 'Password must include at least one small letter.'
    if not any(ch.isdigit() for ch in password):
        return False, 'Password must include at least one number.'
    return True, ''


def get_csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


app.jinja_env.globals['csrf_token'] = get_csrf_token


def too_many_login_attempts(username):
    cutoff = (datetime.now() - timedelta(minutes=LOGIN_LOCK_MINUTES)).strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    row = conn.execute('''SELECT COUNT(*) AS total FROM login_attempts
                          WHERE username=? AND ip_address=? AND success='No' AND attempted_at>=?''',
                       (username, get_client_ip(), cutoff)).fetchone()
    conn.close()
    return row['total'] >= LOGIN_MAX_ATTEMPTS


def record_login_attempt(username, success):
    conn = get_db()
    conn.execute('INSERT INTO login_attempts(username, ip_address, success, attempted_at) VALUES(?,?,?,?)',
                 (username, get_client_ip(), 'Yes' if success else 'No', datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()


def too_many_result_lookup_attempts():
    """Limit repeated public result searches from the same connection."""
    cutoff = (datetime.now() - timedelta(minutes=RESULT_LOOKUP_LOCK_MINUTES)).strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    row = conn.execute('''SELECT COUNT(*) AS total FROM login_attempts
                          WHERE username='pupil-result-lookup' AND ip_address=?
                          AND success='No' AND attempted_at>=?''',
                       (get_client_ip(), cutoff)).fetchone()
    conn.close()
    return row['total'] >= RESULT_LOOKUP_MAX_ATTEMPTS


def record_result_lookup_attempt(success):
    conn = get_db()
    conn.execute('INSERT INTO login_attempts(username, ip_address, success, attempted_at) VALUES(?,?,?,?)',
                 ('pupil-result-lookup', get_client_ip(), 'Yes' if success else 'No',
                  datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()


def normalized_pupil_name(value):
    """Compare registered names without being affected by capital letters or extra spaces."""
    return ' '.join((value or '').casefold().split())


@app.before_request
def validate_csrf():
    if request.method in {'POST', 'PUT', 'PATCH', 'DELETE'}:
        expected = session.get('_csrf_token', '')
        supplied = request.form.get('_csrf_token', '') or request.headers.get('X-CSRF-Token', '')
        if not expected or not supplied or not secrets.compare_digest(expected, supplied):
            abort(400, description='The form expired or could not be verified. Go back, refresh the page and try again.')


@app.before_request
def secure_session_refresh():
    if 'user_id' in session:
        session.permanent = True


@app.before_request
def enforce_required_password_change():
    allowed = {'change_password', 'logout', 'static', 'profile_picture', 'background_picture'}
    if session.get('user_id') and session.get('must_change_password') and request.endpoint not in allowed:
        flash('For security, create your own password before using the portal.', 'warning')
        return redirect(url_for('change_password'))


@app.after_request
def set_security_headers(response):
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'same-origin'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    # Keep this CSP practical for the school app because it uses inline styles/scripts and embedded Google Maps.
    response.headers['Content-Security-Policy'] = "default-src 'self'; img-src 'self' data:; media-src 'self'; frame-src https://www.google.com https://maps.google.com; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline';"
    if IS_PRODUCTION:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    if session.get('user_id') or request.endpoint == 'pupil_results_lookup':
        response.headers['Cache-Control'] = 'no-store, private'
    if request.endpoint == 'pupil_results_lookup':
        response.headers['X-Robots-Tag'] = 'noindex, nofollow'
    return response


@app.errorhandler(400)
def bad_request(error):
    return render_template('error.html', title='Form expired', message=getattr(error, 'description', 'The request could not be verified.')), 400


@app.errorhandler(RequestEntityTooLarge)
def upload_too_large(error):
    return render_template('error.html', title='File too large', message=f'The maximum request size is {MAX_UPLOAD_MB} MB.'), 413


def can_manage_backgrounds():
    return session.get('role') in BACKGROUND_MANAGEMENT_ROLES


def can_manage_calendar():
    return session.get('role') in CALENDAR_MANAGEMENT_ROLES


def student_results_open_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get('role') == 'student' and not is_student_results_active():
            flash('Pupil results portal is currently deactivated. Please wait until school management activates it for results access.', 'warning')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return wrapper

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login first.', 'warning'); return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper


def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if session.get('role') not in roles:
                flash('You are not allowed to access that page.', 'danger'); return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return wrapper
    return decorator


def refresh_session_user(user_id):
    conn = get_db(); user = conn.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone(); conn.close()
    if user:
        session['full_name'] = user['full_name']; session['department'] = user['department']; session['role'] = user['role']


def establish_user_session(user, profile_first=False):
    session.clear()
    session.permanent = True
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['role'] = user['role']
    session['full_name'] = user['full_name']
    session['department'] = user['department']
    session['must_change_password'] = bool(user['must_change_password'])
    if profile_first:
        session['post_password_change_endpoint'] = 'profile'


def sync_staff_return_profile(conn, user_id, full_name, phone, email, address, position, department, qualification, subjects_taught='', updated_at=None):
    """Keep the staff-return identity record aligned with the live staff profile."""
    timestamp = updated_at or datetime.now().strftime('%Y-%m-%d %H:%M')
    conn.execute('''INSERT INTO staff_returns(
        user_id, full_name, phone, email, address, position, department,
        subjects_taught, highest_qualification, submitted_at, updated_at
    ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
    ON CONFLICT(user_id) DO UPDATE SET
        full_name=excluded.full_name,
        phone=excluded.phone,
        email=excluded.email,
        address=excluded.address,
        position=excluded.position,
        department=excluded.department,
        subjects_taught=CASE WHEN excluded.subjects_taught!='' THEN excluded.subjects_taught ELSE staff_returns.subjects_taught END,
        highest_qualification=excluded.highest_qualification,
        updated_at=excluded.updated_at''', (
        user_id, full_name, phone, email, address, position,
        DEPARTMENTS.get(department, department or ''), subjects_taught,
        qualification, timestamp, timestamp
    ))


def calc_result(test1, test2, end_term):
    total = float(test1 or 0) + float(test2 or 0) + float(end_term or 0)
    average = total / 3 if total else 0
    return round(total, 2), round(average, 2)


def log_password_action(conn, target_user, action, note=''):
    """Record every password change/reset so Headteacher and HR can know who changed it."""
    conn.execute('''INSERT INTO password_audit_logs(
                    target_user_id, target_username, target_full_name, target_role, action,
                    changed_by_user_id, changed_by_name, changed_by_role, changed_at, note)
                    VALUES(?,?,?,?,?,?,?,?,?,?)''',
                 (target_user['id'], target_user['username'], target_user['full_name'], target_user['role'], action,
                  session.get('user_id'), session.get('full_name'), session.get('role'),
                  datetime.now().strftime('%Y-%m-%d %H:%M:%S'), note))



def send_reset_email_or_save(conn, user, reset_link):
    """Send recovery by SMTP; only demo mode may display/store a usable reset link."""
    email = (user['email'] or '').strip()
    status = 'Not sent'
    details = 'SMTP is not configured. No usable reset link was stored in production.'
    sent_at = None

    smtp_host = os.environ.get('SMTP_HOST', '').strip()
    smtp_port = int(os.environ.get('SMTP_PORT', '587'))
    smtp_user = os.environ.get('SMTP_USER', '').strip()
    smtp_pass = os.environ.get('SMTP_PASSWORD', '').strip()
    smtp_from = os.environ.get('SMTP_FROM', smtp_user or 'no-reply@kafubublock.edu.zm').strip()

    if smtp_host and smtp_user and smtp_pass:
        try:
            msg = EmailMessage()
            msg['Subject'] = 'Kafubu Block Secondary School portal password recovery'
            msg['From'] = smtp_from
            msg['To'] = email
            body = (
                f"Hello {user['full_name']},\n\n"
                "A password reset was requested for your Kafubu Block Secondary School portal account.\n\n"
                "Click this link to create a new password. The link expires in 30 minutes:\n"
                f"{reset_link}\n\n"
                "If you did not request this reset, ignore this message and inform the system administrator.\n"
            )
            msg.set_content(body)
            with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.send_message(msg)
            status = 'Sent'
            details = 'Reset email sent successfully.'
            sent_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        except Exception as exc:
            status = 'Email error'
            details = f'Email sending failed: {str(exc)[:300]}. No usable link was stored in production.'

    stored_link = reset_link if (status == 'Sent' or DEMO_MODE) else '[not stored]'
    conn.execute('''INSERT INTO password_reset_email_outbox(user_id, username, email, reset_link, status, created_at, sent_at, details)
                    VALUES(?,?,?,?,?,?,?,?)''',
                 (user['id'], user['username'], email, stored_link, status,
                  datetime.now().strftime('%Y-%m-%d %H:%M:%S'), sent_at, details))
    return status



def analyse_results(rows):
    """Return simple result analysis from result rows."""
    total_subjects = len(rows)
    total_test1 = sum(float(r['test1'] or 0) for r in rows)
    total_test2 = sum(float(r['test2'] or 0) for r in rows)
    total_end = sum(float(r['end_term'] or 0) for r in rows)
    total_average = sum(float(r['average'] or 0) for r in rows)
    overall_average = round(total_average / total_subjects, 2) if total_subjects else 0
    best = max(rows, key=lambda r: float(r['average'] or 0), default=None)
    weakest = min(rows, key=lambda r: float(r['average'] or 0), default=None)
    pass_count = sum(1 for r in rows if float(r['average'] or 0) >= 50)
    fail_count = total_subjects - pass_count
    return {
        'total_subjects': total_subjects,
        'average_test1': round(total_test1 / total_subjects, 2) if total_subjects else 0,
        'average_test2': round(total_test2 / total_subjects, 2) if total_subjects else 0,
        'average_end_term': round(total_end / total_subjects, 2) if total_subjects else 0,
        'overall_average': overall_average,
        'best_subject': best['subject'] if best else 'N/A',
        'best_average': round(float(best['average']), 2) if best else 0,
        'weakest_subject': weakest['subject'] if weakest else 'N/A',
        'weakest_average': round(float(weakest['average']), 2) if weakest else 0,
        'pass_count': pass_count,
        'fail_count': fail_count,
        'status': 'Pass' if overall_average >= 50 else ('Needs Improvement' if total_subjects else 'No Results')
    }

def user_can_manage_staff(target_user):
    """Control staff-profile management permissions.
    Headteacher, Deputy Headteacher and HR can manage HOD and Subject Teacher profiles.
    Each HOD can manage only Subject Teacher profiles in his/her own department.
    """
    role = session.get('role')
    target_role = target_user['role']
    if role in STAFF_MANAGEMENT_FULL_ROLES:
        return target_role == 'teacher' or target_role.startswith('hod_')
    if role in ROLE_TO_DEPT:
        return target_role == 'teacher' and target_user['department'] == ROLE_TO_DEPT[role]
    return False


def create_backup_archive(kind='manual'):
    """Create a consistent SQLite snapshot plus uploaded files."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
    filename = f'{kind}_kafubu_backup_{timestamp}.zip'
    final_path = os.path.join(BACKUP_DIR, filename)
    with tempfile.TemporaryDirectory(dir=BACKUP_DIR) as temp_dir:
        snapshot_path = os.path.join(temp_dir, 'school.db')
        source = sqlite3.connect(DB_PATH)
        destination = sqlite3.connect(snapshot_path)
        try:
            source.backup(destination)
        finally:
            destination.close()
            source.close()

        manifest = {
            'application': 'Kafubu Block Secondary School Portal',
            'backup_version': 1,
            'created_at': datetime.now().isoformat(timespec='seconds'),
            'kind': kind,
        }
        temporary_zip = os.path.join(temp_dir, filename)
        with zipfile.ZipFile(temporary_zip, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
            archive.write(snapshot_path, 'school.db')
            archive.writestr('manifest.json', json.dumps(manifest, indent=2))
            for root, _, files in os.walk(UPLOAD_ROOT):
                for upload_name in files:
                    full_path = os.path.join(root, upload_name)
                    archive.write(full_path, os.path.join('uploads', os.path.relpath(full_path, UPLOAD_ROOT)))
        os.replace(temporary_zip, final_path)
    return filename


def list_backup_archives():
    backups = []
    for filename in os.listdir(BACKUP_DIR):
        if not filename.endswith('.zip') or not filename.startswith(('manual_', 'auto_', 'before_restore_')):
            continue
        path = os.path.join(BACKUP_DIR, filename)
        if os.path.isfile(path):
            stat = os.stat(path)
            backups.append({
                'filename': filename,
                'created_at': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                'size_mb': round(stat.st_size / (1024 * 1024), 2),
            })
    return sorted(backups, key=lambda item: item['filename'], reverse=True)


def prune_automatic_backups():
    automatic = [b for b in list_backup_archives() if b['filename'].startswith('auto_')]
    for old in automatic[BACKUP_RETENTION:]:
        os.remove(os.path.join(BACKUP_DIR, old['filename']))


def maybe_create_automatic_backup():
    if not AUTO_BACKUP_ENABLED or not os.path.exists(DB_PATH):
        return
    automatic = [b for b in list_backup_archives() if b['filename'].startswith('auto_')]
    if automatic:
        newest = os.path.join(BACKUP_DIR, automatic[0]['filename'])
        age_hours = (datetime.now().timestamp() - os.path.getmtime(newest)) / 3600
        if age_hours < AUTO_BACKUP_HOURS:
            return
    lock_path = os.path.join(BACKUP_DIR, '.automatic_backup.lock')
    try:
        lock_fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError:
        return
    try:
        os.close(lock_fd)
        create_backup_archive('auto')
        prune_automatic_backups()
    finally:
        if os.path.exists(lock_path):
            os.remove(lock_path)


def restore_backup_archive(filename):
    safe_name = os.path.basename(filename)
    if safe_name != filename:
        raise ValueError('Invalid backup filename.')
    backup_path = os.path.join(BACKUP_DIR, safe_name)
    if not os.path.isfile(backup_path):
        raise ValueError('Backup file was not found.')

    create_backup_archive('before_restore')
    with tempfile.TemporaryDirectory(dir=BACKUP_DIR) as temp_dir:
        with zipfile.ZipFile(backup_path, 'r') as archive:
            names = archive.namelist()
            if 'school.db' not in names or 'manifest.json' not in names:
                raise ValueError('This is not a valid school portal backup.')
            restored_db = os.path.join(temp_dir, 'school.db')
            with archive.open('school.db') as source, open(restored_db, 'wb') as target:
                shutil.copyfileobj(source, target)
            check = sqlite3.connect(restored_db)
            try:
                integrity = check.execute('PRAGMA integrity_check').fetchone()[0]
            finally:
                check.close()
            if integrity != 'ok':
                raise ValueError('The backup database failed its integrity check.')

            os.replace(restored_db, DB_PATH)
            for member in archive.infolist():
                normalized = member.filename.replace('\\', '/')
                if member.is_dir() or not normalized.startswith('uploads/') or '..' in normalized.split('/'):
                    continue
                relative_upload = normalized[len('uploads/'):]
                target_path = os.path.abspath(os.path.join(UPLOAD_ROOT, relative_upload))
                if not target_path.startswith(os.path.abspath(UPLOAD_ROOT) + os.sep):
                    continue
                os.makedirs(os.path.dirname(target_path), exist_ok=True)
                with archive.open(member) as source, open(target_path, 'wb') as target:
                    shutil.copyfileobj(source, target)
    init_db()



@app.route('/')
def index():
    conn = get_db()
    needs = conn.execute("SELECT * FROM donor_needs WHERE status='Needed' ORDER BY priority DESC LIMIT 3").fetchall()
    guidance_posts = conn.execute('SELECT * FROM guidance_posts ORDER BY posted_at DESC LIMIT 3').fetchall()
    conn.close()
    return render_template('index.html', needs=needs, guidance_posts=guidance_posts)
@app.route('/about')
def about():
    return render_template('about.html', content=get_website_content())
@app.route('/departments')
def departments(): return render_template('departments.html')
@app.route('/department/<dept>')
def department_detail(dept):
    if dept not in DEPARTMENTS:
        flash('Department not found.', 'danger')
        return redirect(url_for('departments'))
    conn = get_db()
    materials = conn.execute('SELECT * FROM materials WHERE department=? ORDER BY uploaded_at DESC LIMIT 5', (dept,)).fetchall()
    hod = conn.execute('SELECT * FROM users WHERE department=? AND role LIKE "hod_%" AND is_active=1 LIMIT 1', (dept,)).fetchone()
    teachers = conn.execute('''SELECT * FROM users
                               WHERE department=? AND role='teacher' AND is_active=1
                               ORDER BY full_name''', (dept,)).fetchall()
    conn.close()
    return render_template('department_detail.html', dept=dept, materials=materials, hod=hod, teachers=teachers, teacher_count=len(teachers))
@app.route('/leadership')
def leadership():
    conn=get_db(); leaders=conn.execute("SELECT * FROM users WHERE role IN ('headteacher','deputy_headteacher','hr') OR role LIKE 'hod_%' ORDER BY role").fetchall(); conn.close()
    return render_template('leadership.html', leaders=leaders)

@app.route('/contact-location')
def contact_location():
    return render_template('contact_location.html', content=get_website_content())

@app.route('/edit-contact-location', methods=['GET','POST'])
@login_required
@roles_required('headteacher')
def edit_contact_location():
    keys = ['contact_phone', 'contact_email', 'office_hours', 'school_location', 'map_query', 'contact_message']
    if request.method == 'POST':
        conn = get_db()
        for key in keys:
            value = request.form.get(key, '').strip()
            conn.execute("""INSERT INTO website_content(content_key, content_value, updated_by, updated_at)
                            VALUES(?,?,?,?)
                            ON CONFLICT(content_key) DO UPDATE SET content_value=excluded.content_value, updated_by=excluded.updated_by, updated_at=excluded.updated_at""",
                         (key, value, session.get('full_name','Headteacher'), datetime.now().strftime('%Y-%m-%d %H:%M')))
        conn.commit(); conn.close()
        flash('Contact and location information updated successfully.', 'success')
        return redirect(url_for('contact_location'))
    return render_template('edit_contact_location.html', content=get_website_content())

@app.route('/donor-needs')
def donor_needs():
    conn=get_db(); needs=conn.execute('SELECT * FROM donor_needs ORDER BY CASE priority WHEN "High" THEN 1 WHEN "Medium" THEN 2 ELSE 3 END, created_at DESC').fetchall(); conn.close()
    return render_template('donor_needs.html', needs=needs)
@app.route('/donate', methods=['GET','POST'])
@app.route('/donate/<int:need_id>', methods=['GET','POST'])
def donate(need_id=None):
    conn = get_db()
    selected_need = conn.execute('SELECT * FROM donor_needs WHERE id=?', (need_id,)).fetchone() if need_id else None
    needs = conn.execute('SELECT * FROM donor_needs WHERE status="Needed" ORDER BY CASE priority WHEN "High" THEN 1 WHEN "Medium" THEN 2 ELSE 3 END, item').fetchall()
    if request.method == 'POST':
        form_need_id = request.form.get('need_id') or None
        if form_need_id == 'general':
            form_need_id = None
        donor_name = request.form.get('donor_name','').strip()
        donation_type = request.form.get('donation_type','Money')
        if not donor_name:
            flash('Please enter donor name or organisation name.', 'danger')
        else:
            conn.execute('''INSERT INTO donation_pledges(need_id, donor_name, email, phone, donation_type, amount, item_description, payment_method, message, status, pledged_at)
                            VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
                         (form_need_id, donor_name, request.form.get('email','').strip(), request.form.get('phone','').strip(), donation_type, request.form.get('amount','').strip(), request.form.get('item_description','').strip(), request.form.get('payment_method','').strip(), request.form.get('message','').strip(), 'Pending', datetime.now().strftime('%Y-%m-%d %H:%M')))
            conn.commit(); conn.close()
            flash('Thank you. Your donation pledge has been submitted. The school will contact you using the details provided.', 'success')
            return redirect(url_for('donor_needs'))
    conn.close()
    return render_template('donate.html', needs=needs, selected_need=selected_need)

@app.route('/donation-records')
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def donation_records():
    conn = get_db()
    records = conn.execute('''SELECT d.*, n.item AS need_item FROM donation_pledges d
                              LEFT JOIN donor_needs n ON d.need_id=n.id
                              ORDER BY d.pledged_at DESC''').fetchall()
    conn.close()
    return render_template('donation_records.html', records=records)

@app.route('/update-donation-status/<int:donation_id>', methods=['POST'])
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def update_donation_status(donation_id):
    conn = get_db()
    conn.execute('UPDATE donation_pledges SET status=? WHERE id=?', (request.form.get('status','Pending'), donation_id))
    conn.commit(); conn.close()
    flash('Donation record updated.', 'success')
    return redirect(url_for('donation_records'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        if too_many_login_attempts(username):
            log_security_event('Blocked login attempt', username=username, details='Too many failed login attempts.')
            flash(f'Too many wrong login attempts. Try again after {LOGIN_LOCK_MINUTES} minutes or ask the Headteacher/HR to reset your password.', 'danger')
            return render_template('login.html')
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        conn.close()
        if user and user['is_active'] and check_password_hash(user['password'], password):
            record_login_attempt(username, True)
            establish_user_session(user, profile_first=user['role'] == 'teacher')
            log_security_event('Login successful', username=user['username'], role=user['role'])
            flash(f'Login successful. Welcome to the {user["position"]} portal.', 'success')
            destination = 'profile' if user['role'] == 'teacher' else 'dashboard'
            return redirect(url_for('change_password' if user['must_change_password'] else destination))
        record_login_attempt(username, False)
        log_security_event('Failed login', username=username, details='Wrong username or password.')
        flash('Wrong username or password.', 'danger')
    return render_template('login.html')


@app.route('/teacher-login', methods=['GET', 'POST'])
def teacher_login():
    """Dedicated login for individual Subject Teacher accounts."""
    if session.get('user_id'):
        return redirect(url_for('profile' if session.get('role') == 'teacher' else 'dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        if too_many_login_attempts(username):
            log_security_event('Blocked teacher login attempt', username=username, details='Too many failed login attempts.')
            flash(f'Too many wrong login attempts. Try again after {LOGIN_LOCK_MINUTES} minutes or ask the Headteacher/HR to reset the password.', 'danger')
            return render_template('teacher_login.html')
        conn = get_db()
        teacher = conn.execute("SELECT * FROM users WHERE username=? AND role='teacher'", (username,)).fetchone()
        conn.close()
        if teacher and teacher['is_active'] and check_password_hash(teacher['password'], password):
            record_login_attempt(username, True)
            establish_user_session(teacher, profile_first=True)
            log_security_event('Teacher portal login successful', username=teacher['username'], role=teacher['role'])
            flash(f'Welcome, {teacher["full_name"]}. You can view and update your profile here.', 'success')
            return redirect(url_for('change_password' if teacher['must_change_password'] else 'profile'))
        record_login_attempt(username, False)
        log_security_event('Failed teacher portal login', username=username, details='Wrong teacher username or password.')
        flash('Wrong teacher username or password.', 'danger')
    return render_template('teacher_login.html')


@app.route('/forgot-management-password', methods=['GET','POST'])
def forgot_management_password():
    """Tight recovery: only the approved Google/Gmail address can request Headteacher/HR reset links.
    The user must also provide the correct management username.
    """
    reset_link = None
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        username = request.form.get('username','').strip().lower()
        # Always use a generic public response so attackers cannot confirm valid accounts.
        generic_message = 'If the username and authorised email are correct, a secure reset link will be prepared.'
        if not RECOVERY_ALLOWED_EMAIL or email != RECOVERY_ALLOWED_EMAIL or not username:
            log_security_event('Password recovery blocked', username=username or email, details='Unauthorised email or missing username.')
            flash(generic_message, 'info')
            return render_template('forgot_management_password.html', reset_link=None)

        conn = get_db()
        user = conn.execute('''SELECT * FROM users
                               WHERE lower(username)=?
                               AND lower(email)=?
                               AND role IN ('headteacher','hr')
                               LIMIT 1''', (username, RECOVERY_ALLOWED_EMAIL)).fetchone()
        if not user:
            conn.close()
            log_security_event('Password recovery blocked', username=username, details='Username not Headteacher/HR or email not authorised.')
            flash(generic_message, 'info')
            return render_template('forgot_management_password.html', reset_link=None)

        # Invalidate previous unused links for the same account before creating a new one.
        conn.execute("UPDATE password_reset_tokens SET used='Yes', used_at=? WHERE user_id=? AND used='No'",
                     (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user['id']))
        token = secrets.token_urlsafe(64)
        token_hash = hashlib.sha256(token.encode('utf-8')).hexdigest()
        expires_at = (datetime.now() + timedelta(minutes=15)).strftime('%Y-%m-%d %H:%M:%S')
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute('''INSERT INTO password_reset_tokens(user_id, email, token, role, expires_at, used, created_at)
                        VALUES(?,?,?,?,?,'No',?)''',
                     (user['id'], RECOVERY_ALLOWED_EMAIL, token_hash, user['role'], expires_at, created_at))
        reset_link = url_for('reset_management_password', token=token, _external=True)
        delivery_status = send_reset_email_or_save(conn, user, reset_link)
        log_password_action(conn, user, 'Secure password recovery link requested', 'Recovery allowed only through the authorised Google/Gmail email.')
        conn.commit(); conn.close()
        log_security_event('Secure password recovery link prepared', username=user['username'], role=user['role'])
        if delivery_status == 'Sent':
            flash('A secure password reset link was sent to the authorised management email.', 'success')
        elif DEMO_MODE:
            flash('Demo mode: the reset link is shown below.', 'success')
        else:
            flash('Email delivery is not configured. Ask the hosting administrator to configure SMTP.', 'warning')
        return render_template('forgot_management_password.html', reset_link=reset_link if DEMO_MODE else None)
    return render_template('forgot_management_password.html', reset_link=reset_link)


@app.route('/reset-management-password/<token>', methods=['GET','POST'])
def reset_management_password(token):
    conn = get_db()
    token_hash = hashlib.sha256(token.encode('utf-8')).hexdigest()
    token_row = conn.execute('''SELECT prt.*, u.username, u.full_name, u.password, u.role
                                FROM password_reset_tokens prt
                                JOIN users u ON u.id = prt.user_id
                                WHERE prt.token=? AND prt.used='No'
                                LIMIT 1''', (token_hash,)).fetchone()
    if not token_row:
        conn.close(); flash('Invalid or already used password reset link.', 'danger'); return redirect(url_for('login'))
    if datetime.strptime(token_row['expires_at'], '%Y-%m-%d %H:%M:%S') < datetime.now():
        conn.close(); flash('This password reset link has expired. Please request a new one.', 'warning'); return redirect(url_for('forgot_management_password'))

    if request.method == 'POST':
        new_password = request.form.get('new_password','').strip()
        confirm_password = request.form.get('confirm_password','').strip()
        if not new_password or not confirm_password:
            conn.close(); flash('Enter and confirm the new password.', 'warning'); return redirect(request.url)
        if new_password != confirm_password:
            conn.close(); flash('New password and confirm password do not match.', 'danger'); return redirect(request.url)
        ok, msg = password_is_strong(new_password)
        if not ok:
            conn.close(); flash(msg, 'warning'); return redirect(request.url)
        conn.execute('UPDATE users SET password=?, must_change_password=0, is_active=1 WHERE id=?', (generate_password_hash(new_password), token_row['user_id']))
        conn.execute('UPDATE password_reset_tokens SET used=?, used_at=? WHERE id=?',
                     ('Yes', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), token_row['id']))
        target = conn.execute('SELECT * FROM users WHERE id=?', (token_row['user_id'],)).fetchone()
        log_password_action(conn, target, 'Password reset by email recovery', 'Headteacher/HR reset password using Google email verification link.')
        conn.commit(); conn.close()
        log_security_event('Management password reset completed', username=token_row['username'], role=token_row['role'])
        flash('Password changed successfully. You can now login using the new password.', 'success')
        return redirect(url_for('login'))
    conn.close()
    return render_template('reset_management_password.html', token=token, full_name=token_row['full_name'])


@app.route('/password-reset-email-outbox')
@login_required
@roles_required('headteacher', 'hr')
def password_reset_email_outbox():
    conn = get_db()
    rows = conn.execute('SELECT * FROM password_reset_email_outbox ORDER BY created_at DESC LIMIT 200').fetchall()
    conn.close()
    return render_template('password_reset_email_outbox.html', rows=rows)


@app.route('/teacher-register', methods=['GET','POST'])
@login_required
def teacher_register():
    """Create a teacher account automatically assigned to the registering HOD's department."""
    if not session.get('role','').startswith('hod_'):
        flash('Only HODs can register new Subject Teachers.', 'danger')
        return redirect(url_for('dashboard'))
    hod_department = ROLE_TO_DEPT.get(session.get('role'))
    if request.method == 'POST':
        username = request.form.get('username','').strip().lower()
        password = request.form.get('password','')
        confirm_password = request.form.get('confirm_password','')
        full_name = request.form.get('full_name','').strip()
        # Never trust a submitted department: the HOD's role controls the assignment.
        department = hod_department
        phone = request.form.get('phone','').strip()
        email = request.form.get('email','').strip()
        qualification = request.form.get('qualification','').strip()
        address = request.form.get('address','').strip()
        bio = request.form.get('bio','').strip()
        subjects_taught = request.form.get('subjects_taught','').strip()
        position = request.form.get('position','Subject Teacher').strip() or 'Subject Teacher'
        if not username or not password or not full_name or not department:
            flash('Please fill in username, password, full name and department.', 'danger')
            return render_template('teacher_register.html', hod_department=hod_department, selected_department=department)
        if department not in DEPARTMENTS:
            flash('Please select a valid department.', 'danger')
            return render_template('teacher_register.html', hod_department=hod_department, selected_department=department)
        if password != confirm_password:
            flash('Password and confirm password do not match.', 'danger')
            return render_template('teacher_register.html', hod_department=hod_department, selected_department=department)
        ok, msg = password_is_strong(password)
        if not ok:
            flash(msg, 'warning')
            return render_template('teacher_register.html', hod_department=hod_department, selected_department=department)
        conn = get_db()
        existing = conn.execute('SELECT id FROM users WHERE username=?', (username,)).fetchone()
        if existing:
            conn.close()
            flash('That username is already taken. Choose another username.', 'danger')
            return render_template('teacher_register.html', hod_department=hod_department, selected_department=department)
        filename = None
        file = request.files.get('profile_picture')
        if file and file.filename:
            if not is_allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
                conn.close()
                flash('Only JPG, PNG, GIF or WEBP profile pictures are allowed.', 'danger')
                return render_template('teacher_register.html', hod_department=hod_department, selected_department=department)
            if not uploaded_file_within_limit(file, MAX_IMAGE_BYTES):
                conn.close()
                flash(f'Profile pictures must be {MAX_IMAGE_MB} MB or less.', 'danger')
                return render_template('teacher_register.html', hod_department=hod_department, selected_department=department)
            filename = unique_safe_filename(file.filename)
            file.save(os.path.join(PROFILE_DIR, filename))
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        cur = conn.execute("""INSERT INTO users(username,password,role,full_name,position,department,bio,email,phone,qualification,address,profile_picture,must_change_password,is_active,profile_updated_at)
                              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,1,1,?)""",
                           (username, generate_password_hash(password), 'teacher', full_name, position, department, bio, email, phone, qualification, address, filename, now))
        new_user_id = cur.lastrowid
        sync_staff_return_profile(conn, new_user_id, full_name, phone, email, address, position, department, qualification, subjects_taught, now)
        conn.commit(); conn.close()
        log_security_event('New subject teacher registered', username=username, role='teacher', details=f'Department: {DEPARTMENTS.get(department, department)}')
        log_security_event('New subject teacher credentials issued', username=username, role='teacher', details='Temporary password created by HOD; private password required at first login.')
        return render_template(
            'teacher_registration_complete.html', teacher_name=full_name,
            teacher_username=username, temporary_password=password,
            department_name=DEPARTMENTS.get(department, department), department_key=department
        )
    return render_template('teacher_register.html', hod_department=hod_department, selected_department=hod_department)

@app.route('/logout')
def logout():
    log_security_event('Logout')
    session.clear(); flash('You have logged out.','info'); return redirect(url_for('index'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not user:
        conn.close(); session.clear(); return redirect(url_for('login'))
    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        if not check_password_hash(user['password'], current_password):
            conn.close(); flash('Current or temporary password is incorrect.', 'danger'); return redirect(url_for('change_password'))
        if new_password != confirm_password:
            conn.close(); flash('New password and confirmation do not match.', 'danger'); return redirect(url_for('change_password'))
        ok, message = password_is_strong(new_password)
        if not ok:
            conn.close(); flash(message, 'warning'); return redirect(url_for('change_password'))
        if check_password_hash(user['password'], new_password):
            conn.close(); flash('Choose a password different from the current password.', 'warning'); return redirect(url_for('change_password'))
        conn.execute('UPDATE users SET password=?, must_change_password=0, is_active=1 WHERE id=?',
                     (generate_password_hash(new_password), user['id']))
        log_password_action(conn, user, 'Self password change', 'User created a private portal password.')
        conn.commit(); conn.close()
        session['must_change_password'] = False
        log_security_event('Required password change completed', username=user['username'], role=user['role'])
        flash('Your new private password has been saved.', 'success')
        destination = session.pop('post_password_change_endpoint', 'dashboard')
        if destination not in {'dashboard', 'profile'}:
            destination = 'dashboard'
        return redirect(url_for(destination))
    conn.close()
    return render_template('change_password.html', forced=bool(session.get('must_change_password')))

@app.route('/dashboard')
@login_required
def dashboard():
    conn=get_db()
    material_count=conn.execute('SELECT COUNT(*) AS total FROM materials').fetchone()['total']
    doc_count=conn.execute('SELECT COUNT(*) AS total FROM documents').fetchone()['total']
    need_count=conn.execute('SELECT COUNT(*) AS total FROM donor_needs WHERE status="Needed"').fetchone()['total']
    student_count=conn.execute('SELECT COUNT(*) AS total FROM students').fetchone()['total']
    result_count=conn.execute('SELECT COUNT(*) AS total FROM results').fetchone()['total']
    download_count=conn.execute('SELECT COUNT(*) AS total FROM result_download_logs').fetchone()['total']
    donation_count=conn.execute('SELECT COUNT(*) AS total FROM donation_pledges').fetchone()['total']
    staff_return_count=conn.execute('SELECT COUNT(*) AS total FROM staff_returns').fetchone()['total']
    recent_materials=conn.execute('SELECT * FROM materials ORDER BY uploaded_at DESC LIMIT 6').fetchall()
    recent_docs=conn.execute('SELECT * FROM documents ORDER BY uploaded_at DESC LIMIT 6').fetchall()
    my_results=[]
    if session.get('role') == 'student' and is_student_results_active():
        user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
        student = conn.execute('SELECT * FROM students WHERE student_number=?', (user['student_number'],)).fetchone() if user and user['student_number'] else None
        if student:
            my_results = conn.execute('SELECT * FROM results WHERE student_id=? ORDER BY academic_year DESC, term, subject', (student['id'],)).fetchall()
    conn.close()
    return render_template('dashboard.html', material_count=material_count, doc_count=doc_count, need_count=need_count, student_count=student_count, result_count=result_count, download_count=download_count, donation_count=donation_count, staff_return_count=staff_return_count, recent_materials=recent_materials, recent_docs=recent_docs, my_results=my_results)

@app.route('/profile', methods=['GET','POST'])
@login_required
def profile():
    conn = get_db(); user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if user and user['role'] == 'student':
        conn.close()
        flash('Pupil profile editing has been removed. View your student registration details here.', 'info')
        return redirect(url_for('my_registration'))
    if request.method == 'POST':
        full_name = request.form['full_name']
        position = request.form.get('position') or user['position']
        phone = request.form.get('phone','')
        email = request.form.get('email','')
        bio = request.form.get('bio','')
        address = request.form.get('address','')
        qualification = request.form.get('qualification','')
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        password_blocked_roles = ['teacher', 'guidance_counselling']
        # HOD and Subject Teacher departments are controlled by management assignment.
        if user['role'] in ROLE_TO_DEPT:
            department = ROLE_TO_DEPT[user['role']]
        elif user['role'] == 'teacher':
            department = user['department']
        elif user['role'] == 'student':
            department = user['department']
        else:
            department = request.form.get('department') or None
            if department not in DEPARTMENTS:
                department = None
        filename = user['profile_picture'] if user else None
        file = request.files.get('profile_picture')
        if file and file.filename:
            if not is_allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
                conn.close(); flash('Only JPG, PNG, GIF or WEBP profile pictures are allowed.', 'danger'); return redirect(url_for('profile'))
            if not uploaded_file_within_limit(file, MAX_IMAGE_BYTES):
                conn.close(); flash(f'Profile pictures must be {MAX_IMAGE_MB} MB or less.', 'danger'); return redirect(url_for('profile'))
            filename = unique_safe_filename(file.filename)
            file.save(os.path.join(PROFILE_DIR, filename))
        if user['role'] in password_blocked_roles and (new_password or confirm_password or current_password):
            conn.close(); flash('Subject Teacher password changing has been disabled in the teacher portal. Ask the Headteacher or HR to reset it if needed.', 'warning'); return redirect(url_for('profile'))
        if user['role'] not in ['student'] + password_blocked_roles and (new_password or confirm_password or current_password):
            if not current_password or not new_password or not confirm_password:
                conn.close(); flash('To change your portal password, fill in current password, new password and confirm password.', 'warning'); return redirect(url_for('profile'))
            if not check_password_hash(user['password'], current_password):
                conn.close(); flash('Current password is incorrect.', 'danger'); return redirect(url_for('profile'))
            if new_password != confirm_password:
                conn.close(); flash('New password and confirm password do not match.', 'danger'); return redirect(url_for('profile'))
            ok, msg = password_is_strong(new_password)
            if not ok:
                conn.close(); flash(msg, 'warning'); return redirect(url_for('profile'))
            conn.execute('UPDATE users SET password=?, must_change_password=0 WHERE id=?', (generate_password_hash(new_password), session['user_id']))
            session['must_change_password'] = False
            log_password_action(conn, user, 'Self password change', 'User changed own portal password from My Profile.')
        elif user['role'] == 'student' and (new_password or confirm_password or current_password):
            flash('Pupil password changing has been disabled by the school portal.', 'warning')
        profile_updated_at = datetime.now().strftime('%Y-%m-%d %H:%M')
        conn.execute('''UPDATE users SET full_name=?, position=?, department=?, phone=?, email=?, bio=?, address=?, qualification=?, profile_picture=?, profile_updated_at=? WHERE id=?''',
                     (full_name, position, department, phone, email, bio, address, qualification, filename, profile_updated_at, session['user_id']))
        sync_staff_return_profile(conn, session['user_id'], full_name, phone, email, address, position, department, qualification, updated_at=profile_updated_at)
        conn.commit(); conn.close(); refresh_session_user(session['user_id'])
        if user['role'] == 'student':
            flash('Profile updated successfully. Password changing is disabled for students.','success')
        else:
            
            if user['role'] in password_blocked_roles:
                flash(f'Profile updated successfully. Your details are now updated under {DEPARTMENTS.get(department, "your department")}.','success')
            else:
                flash('Profile and portal password settings updated successfully.','success')
        log_security_event('Staff profile and department record updated', details=f'Department: {DEPARTMENTS.get(department, department or "Not assigned")}')
        return redirect(url_for('profile'))
    conn.close(); return render_template('profile.html', user=user)

@app.route('/profile-picture/<filename>')
def profile_picture(filename): return send_from_directory(PROFILE_DIR, filename)


@app.route('/my-registration')
@login_required
@roles_required('student')
def my_registration():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    student = conn.execute('SELECT * FROM students WHERE student_number=?', (user['student_number'],)).fetchone() if user and user['student_number'] else None
    conn.close()
    return render_template('my_registration.html', student=student, user=user)

@app.route('/materials')
@login_required
def materials():
    dept=request.args.get('dept',''); conn=get_db(); rows=conn.execute('SELECT * FROM materials WHERE department=? ORDER BY uploaded_at DESC',(dept,)).fetchall() if dept in DEPARTMENTS else conn.execute('SELECT * FROM materials ORDER BY uploaded_at DESC').fetchall(); conn.close(); return render_template('materials.html', materials=rows, selected_dept=dept)
@app.route('/upload-material', methods=['GET','POST'])
@login_required
@roles_required(*ROLE_TO_DEPT.keys(), 'headteacher', 'deputy_headteacher', 'hr')
def upload_material():
    allowed_dept=ROLE_TO_DEPT.get(session.get('role'))
    if request.method=='POST':
        title=request.form['title']; description=request.form.get('description',''); department=request.form['department']
        if allowed_dept and department != allowed_dept: flash('HODs can only upload to their own department.','danger'); return redirect(url_for('upload_material'))
        file=request.files.get('file')
        if not file or file.filename=='': flash('Please choose a file.','warning'); return redirect(url_for('upload_material'))
        
        if not is_allowed_file(file.filename, ALLOWED_DOCUMENT_EXTENSIONS):
            flash('This material file type is not allowed. Upload PDF, Word, Excel, PowerPoint, TXT or CSV only.','danger'); return redirect(url_for('upload_material'))
        if not uploaded_file_within_limit(file, MAX_DOCUMENT_BYTES):
            flash(f'Department materials must be {MAX_DOCUMENT_MB} MB or less.','danger'); return redirect(url_for('upload_material'))
        filename=unique_safe_filename(file.filename); file.save(os.path.join(MATERIAL_DIR, filename)); conn=get_db(); conn.execute('INSERT INTO materials(title,description,department,filename,uploaded_by,uploaded_at) VALUES(?,?,?,?,?,?)',(title,description,department,filename,session['full_name'],datetime.now().strftime('%Y-%m-%d %H:%M'))); conn.commit(); conn.close(); flash('Material uploaded successfully.','success'); return redirect(url_for('materials', dept=department))
    return render_template('upload_material.html', allowed_dept=allowed_dept)
@app.route('/official-documents')
@login_required
def official_documents():
    conn=get_db(); docs=conn.execute('SELECT * FROM documents ORDER BY uploaded_at DESC').fetchall(); conn.close(); return render_template('official_documents.html', docs=docs)
@app.route('/upload-document', methods=['GET','POST'])
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def upload_document():
    if request.method=='POST':
        title=request.form['title']; description=request.form.get('description',''); category=request.form['category']; file=request.files.get('file')
        if not file or file.filename=='': flash('Please choose a file.','warning'); return redirect(url_for('upload_document'))
        
        if not is_allowed_file(file.filename, ALLOWED_DOCUMENT_EXTENSIONS):
            flash('This official document file type is not allowed. Upload PDF, Word, Excel, PowerPoint, TXT or CSV only.','danger'); return redirect(url_for('upload_document'))
        if not uploaded_file_within_limit(file, MAX_DOCUMENT_BYTES):
            flash(f'Official documents must be {MAX_DOCUMENT_MB} MB or less.','danger'); return redirect(url_for('upload_document'))
        filename=unique_safe_filename(file.filename); file.save(os.path.join(DOCUMENT_DIR, filename)); conn=get_db(); conn.execute('INSERT INTO documents(title,description,category,filename,uploaded_by,uploaded_at) VALUES(?,?,?,?,?,?)',(title,description,category,filename,session['full_name'],datetime.now().strftime('%Y-%m-%d %H:%M'))); conn.commit(); conn.close(); flash('Official document uploaded successfully.','success'); return redirect(url_for('official_documents'))
    return render_template('upload_document.html')
@app.route('/manage-donor-needs', methods=['GET','POST'])
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def manage_donor_needs():
    if request.method=='POST':
        conn=get_db()
        conn.execute('INSERT INTO donor_needs(item,description,quantity,priority,estimated_cost,status,created_at) VALUES(?,?,?,?,?,?,?)',
                     (request.form['item'], request.form.get('description',''), request.form.get('quantity',''), request.form['priority'], request.form.get('estimated_cost',''), request.form.get('status','Needed'), datetime.now().strftime('%Y-%m-%d %H:%M')))
        conn.commit(); conn.close()
        flash('Need added successfully.','success')
        return redirect(url_for('manage_donor_needs'))
    conn=get_db()
    needs=conn.execute('SELECT * FROM donor_needs ORDER BY created_at DESC').fetchall()
    conn.close()
    return render_template('manage_donor_needs.html', needs=needs)

@app.route('/edit-donor-need/<int:need_id>', methods=['GET','POST'])
@login_required
@roles_required('headteacher')
def edit_donor_need(need_id):
    conn = get_db()
    need = conn.execute('SELECT * FROM donor_needs WHERE id=?', (need_id,)).fetchone()
    if not need:
        conn.close(); flash('Donor need not found.', 'danger'); return redirect(url_for('manage_donor_needs'))
    if request.method == 'POST':
        conn.execute("""UPDATE donor_needs SET item=?, description=?, quantity=?, priority=?, estimated_cost=?, status=? WHERE id=?""",
                     (request.form['item'], request.form.get('description',''), request.form.get('quantity',''), request.form['priority'], request.form.get('estimated_cost',''), request.form.get('status','Needed'), need_id))
        conn.commit(); conn.close()
        flash('Donor need updated successfully.', 'success')
        return redirect(url_for('manage_donor_needs'))
    conn.close()
    return render_template('edit_donor_need.html', need=need)

@app.route('/delete-donor-need/<int:need_id>', methods=['POST'])
@login_required
@roles_required('headteacher')
def delete_donor_need(need_id):
    conn = get_db()
    conn.execute('DELETE FROM donor_needs WHERE id=?', (need_id,))
    conn.commit(); conn.close()
    flash('Donor need deleted successfully.', 'info')
    return redirect(url_for('manage_donor_needs'))

@app.route('/edit-school-about', methods=['GET','POST'])
@login_required
@roles_required('headteacher')
def edit_school_about():
    keys = ['about_intro', 'vision', 'mission', 'core_values', 'motto']
    if request.method == 'POST':
        conn = get_db()
        for key in keys:
            value = request.form.get(key, '').strip()
            conn.execute("""INSERT INTO website_content(content_key, content_value, updated_by, updated_at)
                            VALUES(?,?,?,?)
                            ON CONFLICT(content_key) DO UPDATE SET content_value=excluded.content_value, updated_by=excluded.updated_by, updated_at=excluded.updated_at""",
                         (key, value, session.get('full_name','Headteacher'), datetime.now().strftime('%Y-%m-%d %H:%M')))
        conn.commit(); conn.close()
        flash('About page information updated successfully.', 'success')
        return redirect(url_for('about'))
    return render_template('edit_school_about.html', content=get_website_content())

@app.route('/students', methods=['GET','POST'])
@login_required
@roles_required('teacher', *ROLE_TO_DEPT.keys(), 'headteacher', 'deputy_headteacher', 'hr')
def students():
    if request.method == 'POST':
        conn=get_db(); student_number=request.form['student_number'].strip(); full_name=request.form['full_name'].strip(); class_teacher=request.form.get('class_teacher','').strip(); conn.execute('INSERT INTO students(student_number,full_name,grade,class_name,gender,parent_phone,class_teacher,created_at) VALUES(?,?,?,?,?,?,?,?)', (student_number, full_name, request.form['grade'], request.form['class_name'], request.form.get('gender',''), request.form.get('parent_phone',''), class_teacher, datetime.now().strftime('%Y-%m-%d %H:%M')));
        username=student_number.lower().replace('-', '')
        temporary_password = generate_temporary_password()
        if not conn.execute('SELECT id FROM users WHERE username=?', (username,)).fetchone():
            conn.execute('INSERT INTO users(username,password,role,full_name,position,department,bio,email,phone,student_number,must_change_password,is_active) VALUES(?,?,?,?,?,?,?,?,?,?,1,1)', (username, generate_password_hash(temporary_password), 'student', full_name, 'Pupil', None, 'Pupil portal account for viewing academic results.', '', request.form.get('parent_phone',''), student_number))
        conn.commit(); conn.close(); flash(f'Pupil added. Username: {username}. Temporary password: {temporary_password}. Give it privately to the pupil; it must be changed at first login.', 'success'); return redirect(url_for('students'))
    conn=get_db(); rows=conn.execute('SELECT * FROM students ORDER BY grade, class_name, full_name').fetchall(); conn.close(); return render_template('students.html', students=rows)

@app.route('/delete-student/<int:student_id>', methods=['POST'])
@login_required
@roles_required('teacher', *ROLE_TO_DEPT.keys(), 'headteacher', 'deputy_headteacher', 'hr')
def delete_student(student_id):
    conn = get_db()
    student = conn.execute('SELECT * FROM students WHERE id=?', (student_id,)).fetchone()
    if not student:
        conn.close(); flash('Pupil record not found.', 'danger'); return redirect(url_for('students'))
    conn.execute('DELETE FROM results WHERE student_id=?', (student_id,))
    conn.execute('DELETE FROM users WHERE role="student" AND student_number=?', (student['student_number'],))
    conn.execute('DELETE FROM students WHERE id=?', (student_id,))
    conn.commit(); conn.close()
    flash('Pupil profile, login account and related results were deleted successfully.', 'info')
    return redirect(url_for('students'))



@app.route('/student-records')
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def student_records():
    """HR and school management can view the full student register, totals and summaries."""
    conn = get_db()
    selected_grade = request.args.get('grade','')
    selected_class = request.args.get('class_name','')

    query = 'SELECT * FROM students WHERE 1=1'
    params = []
    if selected_grade:
        query += ' AND grade=?'; params.append(selected_grade)
    if selected_class:
        query += ' AND class_name=?'; params.append(selected_class)
    query += ' ORDER BY grade, class_name, full_name'
    students_rows = conn.execute(query, params).fetchall()

    total_students = conn.execute('SELECT COUNT(*) AS total FROM students').fetchone()['total']
    filtered_total = len(students_rows)
    grade_summary = conn.execute('SELECT grade, COUNT(*) AS total FROM students GROUP BY grade ORDER BY grade').fetchall()
    class_summary = conn.execute('SELECT grade, class_name, COUNT(*) AS total FROM students GROUP BY grade, class_name ORDER BY grade, class_name').fetchall()
    gender_summary = conn.execute('SELECT COALESCE(NULLIF(gender, ""), "Not indicated") AS gender, COUNT(*) AS total FROM students GROUP BY COALESCE(NULLIF(gender, ""), "Not indicated") ORDER BY gender').fetchall()
    classes = conn.execute('SELECT DISTINCT class_name FROM students WHERE class_name IS NOT NULL AND class_name != "" ORDER BY class_name').fetchall()
    conn.close()
    return render_template('student_records.html', students=students_rows, total_students=total_students, filtered_total=filtered_total, grade_summary=grade_summary, class_summary=class_summary, gender_summary=gender_summary, selected_grade=selected_grade, selected_class=selected_class, classes=classes)

@app.route('/download-student-records-excel')
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def download_student_records_excel():
    """Download registered student information and totals in Excel for HR records."""
    selected_grade = request.args.get('grade','')
    selected_class = request.args.get('class_name','')
    conn = get_db()
    query = 'SELECT * FROM students WHERE 1=1'
    params = []
    if selected_grade:
        query += ' AND grade=?'; params.append(selected_grade)
    if selected_class:
        query += ' AND class_name=?'; params.append(selected_class)
    query += ' ORDER BY grade, class_name, full_name'
    rows = conn.execute(query, params).fetchall()
    total_students = conn.execute('SELECT COUNT(*) AS total FROM students').fetchone()['total']
    grade_summary = conn.execute('SELECT grade, COUNT(*) AS total FROM students GROUP BY grade ORDER BY grade').fetchall()
    class_summary = conn.execute('SELECT grade, class_name, COUNT(*) AS total FROM students GROUP BY grade, class_name ORDER BY grade, class_name').fetchall()
    gender_summary = conn.execute('SELECT COALESCE(NULLIF(gender, ""), "Not indicated") AS gender, COUNT(*) AS total FROM students GROUP BY COALESCE(NULLIF(gender, ""), "Not indicated") ORDER BY gender').fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pupil Records'
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    title_font = Font(size=14, bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws['A1'] = 'KAFUBU BLOCK SECONDARY SCHOOL - STUDENT REGISTER'
    ws['A1'].font = title_font
    ws['A2'] = 'Total Pupils at School'
    ws['B2'] = total_students
    ws['A3'] = 'Filtered Records Downloaded'
    ws['B3'] = len(rows)
    ws['A4'] = 'Generated By'
    ws['B4'] = session.get('full_name','')
    ws['A5'] = 'Generated On'
    ws['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    headers = ['Pupil Number','Full Name','Grade/Form','Class','Class Teacher','Gender','Parent/Guardian Phone','Registration Date']
    start_row = 7
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center'); cell.border = border
    for r, student in enumerate(rows, start_row+1):
        values = [student['student_number'], student['full_name'], student['grade'], student['class_name'], student['class_teacher'], student['gender'], student['parent_phone'], student['created_at']]
        for c, value in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top')

    for i, width in enumerate([18, 28, 16, 12, 24, 12, 22, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    summary = wb.create_sheet('Summary')
    summary['A1'] = 'STUDENT TOTALS SUMMARY'
    summary['A1'].font = title_font
    summary['A3'] = 'Total Pupils at School'
    summary['B3'] = total_students

    row = 5
    summary.cell(row=row, column=1, value='Total by Grade/Form').font = Font(bold=True)
    row += 1
    summary.cell(row=row, column=1, value='Grade/Form').fill = header_fill
    summary.cell(row=row, column=2, value='Total').fill = header_fill
    summary.cell(row=row, column=1).font = header_font
    summary.cell(row=row, column=2).font = header_font
    for item in grade_summary:
        row += 1
        summary.cell(row=row, column=1, value=item['grade'])
        summary.cell(row=row, column=2, value=item['total'])

    row += 3
    summary.cell(row=row, column=1, value='Total by Class').font = Font(bold=True)
    row += 1
    for col, h in enumerate(['Grade/Form','Class','Total'], 1):
        cell = summary.cell(row=row, column=col, value=h); cell.fill = header_fill; cell.font = header_font
    for item in class_summary:
        row += 1
        summary.cell(row=row, column=1, value=item['grade'])
        summary.cell(row=row, column=2, value=item['class_name'])
        summary.cell(row=row, column=3, value=item['total'])

    row += 3
    summary.cell(row=row, column=1, value='Total by Gender').font = Font(bold=True)
    row += 1
    for col, h in enumerate(['Gender','Total'], 1):
        cell = summary.cell(row=row, column=col, value=h); cell.fill = header_fill; cell.font = header_font
    for item in gender_summary:
        row += 1
        summary.cell(row=row, column=1, value=item['gender'])
        summary.cell(row=row, column=2, value=item['total'])

    for wsx in [summary]:
        for col in range(1, 5):
            wsx.column_dimensions[get_column_letter(col)].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='student_registered_records.xlsx')

@app.route('/results', methods=['GET','POST'])
@login_required
@roles_required('teacher', *ROLE_TO_DEPT.keys(), 'headteacher', 'deputy_headteacher', 'hr')
def results():
    conn=get_db()
    if request.method == 'POST':
        student_id = request.form['student_id']; subject = request.form['subject']; grade = request.form['grade']; term = request.form['term']; year = request.form['academic_year']
        test1 = float(request.form.get('test1') or 0); test2 = float(request.form.get('test2') or 0); end_term = float(request.form.get('end_term') or 0)
        total, average = calc_result(test1, test2, end_term)
        comment = request.form.get('comment','')
        conn.execute('''INSERT INTO results(student_id,subject,grade,term,academic_year,test1,test2,end_term,total,average,comment,entered_by,entered_at)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''', (student_id, subject, grade, term, year, test1, test2, end_term, total, average, comment, session['full_name'], datetime.now().strftime('%Y-%m-%d %H:%M')))
        conn.commit(); conn.close(); flash('Result added successfully.','success'); return redirect(url_for('results'))
    selected_grade=request.args.get('grade','')
    students_list=conn.execute('SELECT * FROM students ORDER BY grade, class_name, full_name').fetchall()
    if selected_grade:
        rows=conn.execute('''SELECT results.*, students.full_name, students.student_number, students.class_name FROM results JOIN students ON results.student_id=students.id WHERE results.grade=? ORDER BY results.entered_at DESC''', (selected_grade,)).fetchall()
    else:
        rows=conn.execute('''SELECT results.*, students.full_name, students.student_number, students.class_name FROM results JOIN students ON results.student_id=students.id ORDER BY results.entered_at DESC''').fetchall()
    conn.close(); return render_template('results.html', students=students_list, results=rows, selected_grade=selected_grade)

@app.route('/student-results')
@login_required
@student_results_open_required
def student_results():
    conn=get_db(); selected_student_id=request.args.get('student_id')
    if session.get('role') == 'student' and is_student_results_active():
        user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
        student = conn.execute('SELECT * FROM students WHERE student_number=?', (user['student_number'],)).fetchone() if user and user['student_number'] else None
        rows = conn.execute('SELECT * FROM results WHERE student_id=? ORDER BY academic_year DESC, term, subject', (student['id'],)).fetchall() if student else []
        conn.close(); return render_template('student_results.html', student=student, results=rows, students=[])
    if session.get('role') in ['teacher','headteacher','deputy_headteacher','hr'] or session.get('role','').startswith('hod_'):
        students_list=conn.execute('SELECT * FROM students ORDER BY grade, class_name, full_name').fetchall()
        student = conn.execute('SELECT * FROM students WHERE id=?', (selected_student_id,)).fetchone() if selected_student_id else None
        rows = conn.execute('SELECT * FROM results WHERE student_id=? ORDER BY academic_year DESC, term, subject', (selected_student_id,)).fetchall() if selected_student_id else []
        conn.close(); return render_template('student_results.html', student=student, results=rows, students=students_list)
    conn.close(); flash('You are not allowed to access that page.', 'danger'); return redirect(url_for('dashboard'))


@app.route('/pupil-results', methods=['GET', 'POST'])
def pupil_results_lookup():
    """Allow a pupil to view results using the registered number and full name."""
    portal_open = is_student_results_active()
    student = None
    rows = []

    if request.method == 'POST':
        if not portal_open:
            flash('The pupil results portal is currently closed. Please check again after school management activates it.', 'warning')
            return render_template('pupil_results_lookup.html', portal_open=False, student=None, results=[])

        student_number = request.form.get('student_number', '').strip()[:40]
        full_name = request.form.get('full_name', '').strip()[:150]

        if too_many_result_lookup_attempts():
            flash(f'Too many unsuccessful searches. Please wait {RESULT_LOOKUP_LOCK_MINUTES} minutes before trying again.', 'danger')
            return render_template('pupil_results_lookup.html', portal_open=True, student=None, results=[])

        if not student_number or not full_name:
            record_result_lookup_attempt(False)
            flash('Enter both your student number and your full registered name.', 'warning')
            return render_template('pupil_results_lookup.html', portal_open=True, student=None, results=[])

        conn = get_db()
        candidate = conn.execute('SELECT * FROM students WHERE lower(trim(student_number))=lower(?) LIMIT 1',
                                 (student_number,)).fetchone()
        name_matches = bool(candidate and normalized_pupil_name(candidate['full_name']) == normalized_pupil_name(full_name))

        if name_matches:
            student = candidate
            rows = conn.execute('''SELECT * FROM results WHERE student_id=?
                                   ORDER BY academic_year DESC, term, subject''',
                                (student['id'],)).fetchall()
        conn.close()
        record_result_lookup_attempt(name_matches)

        if not name_matches:
            log_security_event('Unsuccessful pupil result lookup', details='The supplied pupil number and name did not match.')
            flash('The student number and name did not match a registered pupil. Check both entries and try again.', 'danger')
        else:
            log_security_event('Pupil result lookup successful', username=student['student_number'], role='student')

    return render_template('pupil_results_lookup.html', portal_open=portal_open, student=student, results=rows)

@app.route('/result-analysis')
@login_required
@student_results_open_required
def result_analysis():
    conn = get_db()
    selected_student_id = request.args.get('student_id','')
    selected_grade = request.args.get('grade','')
    selected_term = request.args.get('term','')
    selected_year = request.args.get('academic_year','')

    if session.get('role') == 'student' and is_student_results_active():
        user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
        student = conn.execute('SELECT * FROM students WHERE student_number=?', (user['student_number'],)).fetchone() if user and user['student_number'] else None
        rows = conn.execute('SELECT * FROM results WHERE student_id=? ORDER BY academic_year DESC, term, subject', (student['id'],)).fetchall() if student else []
        analysis = analyse_results(rows)
        conn.close()
        return render_template('result_analysis.html', student=student, rows=rows, analysis=analysis, students=[], selected_student_id='', selected_grade='', selected_term='', selected_year='', subject_summary=[], grade_summary=[])

    if session.get('role') not in ['teacher','headteacher','deputy_headteacher','hr'] and not session.get('role','').startswith('hod_'):
        conn.close(); flash('You are not allowed to access result analysis.', 'danger'); return redirect(url_for('dashboard'))

    students_list = conn.execute('SELECT * FROM students ORDER BY grade, class_name, full_name').fetchall()
    query = """SELECT results.*, students.full_name, students.student_number, students.class_name
               FROM results JOIN students ON results.student_id=students.id WHERE 1=1"""
    params = []
    if selected_student_id:
        query += ' AND results.student_id=?'; params.append(selected_student_id)
    if selected_grade:
        query += ' AND results.grade=?'; params.append(selected_grade)
    if selected_term:
        query += ' AND results.term=?'; params.append(selected_term)
    if selected_year:
        query += ' AND results.academic_year=?'; params.append(selected_year)
    query += ' ORDER BY results.academic_year DESC, results.term, students.full_name, results.subject'
    rows = conn.execute(query, params).fetchall()
    selected_student = conn.execute('SELECT * FROM students WHERE id=?', (selected_student_id,)).fetchone() if selected_student_id else None
    analysis = analyse_results(rows)
    subject_summary = conn.execute('''SELECT subject, COUNT(*) AS entries, ROUND(AVG(average),2) AS avg_mark, ROUND(AVG(test1),2) AS avg_test1, ROUND(AVG(test2),2) AS avg_test2, ROUND(AVG(end_term),2) AS avg_end_term
                                      FROM results GROUP BY subject ORDER BY subject''').fetchall()
    grade_summary = conn.execute('''SELECT grade, COUNT(*) AS entries, ROUND(AVG(average),2) AS avg_mark
                                    FROM results GROUP BY grade ORDER BY grade''').fetchall()
    conn.close()
    return render_template('result_analysis.html', student=selected_student, rows=rows, analysis=analysis, students=students_list, selected_student_id=selected_student_id, selected_grade=selected_grade, selected_term=selected_term, selected_year=selected_year, subject_summary=subject_summary, grade_summary=grade_summary)

@app.route('/download-result-analysis')
@login_required
def download_result_analysis():
    if session.get('role') not in FULL_ACCESS_ROLES and session.get('role') not in ROLE_TO_DEPT and session.get('role') != 'deputy_headteacher':
        flash('Only HODs, Headteacher, Deputy Headteacher and HR can download result analysis.', 'danger')
        return redirect(url_for('result_analysis'))
    selected_grade = request.args.get('grade','')
    selected_term = request.args.get('term','')
    selected_year = request.args.get('academic_year','')
    conn = get_db()
    query = """SELECT students.student_number, students.full_name, students.class_name, results.grade, results.subject, results.term, results.academic_year, results.test1, results.test2, results.end_term, results.total, results.average, results.comment, results.entered_by
               FROM results JOIN students ON results.student_id=students.id WHERE 1=1"""
    params=[]
    if selected_grade:
        query += ' AND results.grade=?'; params.append(selected_grade)
    if selected_term:
        query += ' AND results.term=?'; params.append(selected_term)
    if selected_year:
        query += ' AND results.academic_year=?'; params.append(selected_year)
    query += ' ORDER BY results.grade, students.full_name, results.subject'
    rows = conn.execute(query, params).fetchall()
    conn.close()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Pupil Number','Pupil Name','Class','Grade','Subject','Term','Academic Year','Test 1','Test 2','End Term','Total','Average','Comment','Entered By'])
    for r in rows:
        writer.writerow([r['student_number'], r['full_name'], r['class_name'], r['grade'], r['subject'], r['term'], r['academic_year'], r['test1'], r['test2'], r['end_term'], r['total'], r['average'], r['comment'], r['entered_by']])
    data = BytesIO(output.getvalue().encode('utf-8'))
    data.seek(0)
    return send_file(data, mimetype='text/csv', as_attachment=True, download_name='student_result_analysis.csv')


@app.route('/student-portal-control', methods=['GET','POST'])
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def student_portal_control():
    conn = get_db()
    if request.method == 'POST':
        status = request.form.get('status', 'inactive')
        if status not in ['active', 'inactive']:
            status = 'inactive'
        conn.execute('''INSERT INTO portal_settings(setting_key, setting_value, updated_by, updated_at)
                        VALUES(?,?,?,?)
                        ON CONFLICT(setting_key) DO UPDATE SET setting_value=excluded.setting_value, updated_by=excluded.updated_by, updated_at=excluded.updated_at''',
                     ('student_results_active', status, session.get('full_name'), datetime.now().strftime('%Y-%m-%d %H:%M')))
        conn.commit()
        flash('Pupil results portal has been ' + ('activated.' if status == 'active' else 'deactivated.'), 'success')
        conn.close()
        return redirect(url_for('student_portal_control'))
    setting = conn.execute('SELECT * FROM portal_settings WHERE setting_key=?', ('student_results_active',)).fetchone()
    conn.close()
    return render_template('student_portal_control.html', setting=setting)

@app.route('/download-my-result-pdf')
@login_required
@student_results_open_required
def download_my_result_pdf():
    conn = get_db()
    if session.get('role') == 'student':
        user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
        student = conn.execute('SELECT * FROM students WHERE student_number=?', (user['student_number'],)).fetchone() if user and user['student_number'] else None
    else:
        student_id = request.args.get('student_id')
        if session.get('role') not in ['teacher','headteacher','deputy_headteacher','hr'] and not session.get('role','').startswith('hod_'):
            conn.close(); flash('You are not allowed to download this report.', 'danger'); return redirect(url_for('dashboard'))
        student = conn.execute('SELECT * FROM students WHERE id=?', (student_id,)).fetchone() if student_id else None
    if not student:
        conn.close(); flash('Pupil record not found.', 'danger'); return redirect(url_for('student_results'))
    rows = conn.execute('SELECT * FROM results WHERE student_id=? ORDER BY academic_year DESC, term, subject', (student['id'],)).fetchall()
    analysis = analyse_results(rows)
    conn.execute('''INSERT INTO result_download_logs(student_id, student_number, student_name, grade, class_name, downloaded_by, downloader_role, downloaded_at)
                    VALUES(?,?,?,?,?,?,?,?)''',
                 (student['id'], student['student_number'], student['full_name'], student['grade'], student['class_name'], session.get('full_name'), session.get('role'), datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.3*cm, leftMargin=1.3*cm, topMargin=1.2*cm, bottomMargin=1.2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleCenter', parent=styles['Title'], alignment=1, fontSize=16, leading=20)
    normal = styles['Normal']
    elements = []
    elements.append(Paragraph('KAFUBU BLOCK SECONDARY SCHOOL', title_style))
    elements.append(Paragraph('STUDENT ACADEMIC REPORT FORM', ParagraphStyle('SubTitle', parent=styles['Heading2'], alignment=1, fontSize=12)))
    elements.append(Spacer(1, 0.25*cm))
    info = [
        ['Pupil Name', student['full_name'], 'Pupil No.', student['student_number']],
        ['Grade/Form', student['grade'], 'Class', student['class_name']],
        ['Class Teacher', student['class_teacher'] or 'Not assigned', 'Generated On', datetime.now().strftime('%Y-%m-%d %H:%M')]
    ]
    t = Table(info, colWidths=[3.2*cm, 5.2*cm, 3.2*cm, 5.2*cm])
    t.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.4,colors.grey),('BACKGROUND',(0,0),(0,-1),colors.lightgrey),('BACKGROUND',(2,0),(2,-1),colors.lightgrey),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),('FONTNAME',(2,0),(2,-1),'Helvetica-Bold')]))
    elements.append(t)
    elements.append(Spacer(1, 0.4*cm))
    data = [['Subject','Term','Year','Test 1','Test 2','End Term','Total','Average','Comment']]
    for r in rows:
        data.append([r['subject'], r['term'], r['academic_year'], str(r['test1']), str(r['test2']), str(r['end_term']), str(r['total']), str(r['average']), Paragraph(r['comment'] or '', normal)])
    if len(data) == 1:
        data.append(['No results uploaded yet.','','','','','','','',''])
    table = Table(data, repeatRows=1, colWidths=[2.5*cm, 1.7*cm, 1.7*cm, 1.5*cm, 1.5*cm, 1.8*cm, 1.5*cm, 1.7*cm, 3.8*cm])
    table.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.35,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#d9eaf7')),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('ALIGN',(3,1),(7,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'TOP'),('FONTSIZE',(0,0),(-1,-1),8)]))
    elements.append(table)
    elements.append(Spacer(1, 0.4*cm))
    summary = [['Result Analysis','Value'], ['Overall Average', f"{analysis['overall_average']}%"], ['Best Subject', f"{analysis['best_subject']} ({analysis['best_average']}%)"], ['Weakest Subject', f"{analysis['weakest_subject']} ({analysis['weakest_average']}%)"], ['Status', analysis['status']]]
    st = Table(summary, colWidths=[5*cm, 8*cm])
    st.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.35,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.lightgrey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')]))
    elements.append(st)
    elements.append(Spacer(1, 0.7*cm))
    elements.append(Paragraph('Class Teacher Signature: ____________________________', normal))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(Paragraph('Headteacher Signature: _____________________________', normal))
    doc.build(elements)
    buffer.seek(0)
    safe_name = secure_filename(student['full_name'].replace(' ', '_')) or 'student'
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'{safe_name}_result_report.pdf')

@app.route('/result-download-logs')
@login_required
@roles_required('teacher', *ROLE_TO_DEPT.keys(), 'headteacher', 'deputy_headteacher', 'hr')
def result_download_logs():
    selected_grade = request.args.get('grade','')
    conn = get_db()
    if selected_grade:
        logs = conn.execute('SELECT * FROM result_download_logs WHERE grade=? ORDER BY downloaded_at DESC', (selected_grade,)).fetchall()
    else:
        logs = conn.execute('SELECT * FROM result_download_logs ORDER BY downloaded_at DESC').fetchall()
    conn.close()
    return render_template('result_download_logs.html', logs=logs, selected_grade=selected_grade)

@app.route('/download-result-download-logs')
@login_required
@roles_required('teacher', *ROLE_TO_DEPT.keys(), 'headteacher', 'deputy_headteacher', 'hr')
def download_result_download_logs():
    selected_grade = request.args.get('grade','')
    conn = get_db()
    if selected_grade:
        logs = conn.execute('SELECT * FROM result_download_logs WHERE grade=? ORDER BY downloaded_at DESC', (selected_grade,)).fetchall()
    else:
        logs = conn.execute('SELECT * FROM result_download_logs ORDER BY downloaded_at DESC').fetchall()
    conn.close()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Pupil Number','Pupil Name','Grade/Form','Class','Downloaded By','Downloader Role','Downloaded At'])
    for log in logs:
        writer.writerow([log['student_number'], log['student_name'], log['grade'], log['class_name'], log['downloaded_by'], log['downloader_role'], log['downloaded_at']])
    data = BytesIO(output.getvalue().encode('utf-8'))
    data.seek(0)
    return send_file(data, mimetype='text/csv', as_attachment=True, download_name='student_result_download_records.csv')

@app.route('/delete-result/<int:result_id>', methods=['POST'])
@login_required
@roles_required('teacher', *ROLE_TO_DEPT.keys(), 'headteacher', 'deputy_headteacher', 'hr')
def delete_result(result_id):
    conn=get_db(); conn.execute('DELETE FROM results WHERE id=?', (result_id,)); conn.commit(); conn.close(); flash('Result deleted.','info'); return redirect(url_for('results'))



def get_staff_return_fields(conn, active_only=True):
    if active_only:
        return conn.execute("SELECT * FROM staff_return_fields WHERE is_active='Yes' ORDER BY display_order, id").fetchall()
    return conn.execute("SELECT * FROM staff_return_fields ORDER BY display_order, id").fetchall()

def get_staff_return_core_fields(conn, active_only=True):
    if active_only:
        return conn.execute("SELECT * FROM staff_return_core_fields WHERE is_active='Yes' ORDER BY display_order, id").fetchall()
    return conn.execute("SELECT * FROM staff_return_core_fields ORDER BY display_order, id").fetchall()

def build_core_field_helpers(core_fields):
    keys = [f['field_key'] for f in core_fields]
    labels = {f['field_key']: f['field_label'] for f in core_fields}
    required = {f['field_key']: (f['is_required'] == 'Yes') for f in core_fields}
    sections = {}
    for f in core_fields:
        sections.setdefault(f['section'], []).append(f['field_key'])
    return keys, labels, required, sections

def get_staff_return_answers(conn, staff_return_id):
    if not staff_return_id:
        return {}
    rows = conn.execute('SELECT field_id, answer_value FROM staff_return_answers WHERE staff_return_id=?', (staff_return_id,)).fetchall()
    return {row['field_id']: row['answer_value'] for row in rows}

def save_staff_return_answers(conn, staff_return_id, custom_fields, form):
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    for field in custom_fields:
        value = form.get(f"custom_{field['id']}", '').strip()
        conn.execute("""INSERT INTO staff_return_answers(staff_return_id, field_id, answer_value, updated_at)
                        VALUES(?,?,?,?)
                        ON CONFLICT(staff_return_id, field_id) DO UPDATE SET answer_value=excluded.answer_value, updated_at=excluded.updated_at""",
                     (staff_return_id, field['id'], value, now))

def fetch_staff_return_export_data(conn, selected_department=''):
    if selected_department:
        rows = conn.execute('SELECT * FROM staff_returns WHERE department=? ORDER BY full_name', (selected_department,)).fetchall()
    else:
        rows = conn.execute('SELECT * FROM staff_returns ORDER BY department, full_name').fetchall()
    custom_fields = get_staff_return_fields(conn, active_only=False)
    answers = {}
    if rows:
        ids = [str(r['id']) for r in rows]
        placeholders = ','.join('?' for _ in ids)
        ans_rows = conn.execute(f'SELECT staff_return_id, field_id, answer_value FROM staff_return_answers WHERE staff_return_id IN ({placeholders})', ids).fetchall()
        for a in ans_rows:
            answers.setdefault(a['staff_return_id'], {})[a['field_id']] = a['answer_value']
    return rows, custom_fields, answers


@app.route('/staff-return', methods=['GET','POST'])
@login_required
def staff_return():
    if session.get('role') == 'student':
        flash('Pupils cannot access staff return registration.', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    existing = conn.execute('SELECT * FROM staff_returns WHERE user_id=?', (session['user_id'],)).fetchone()
    core_fields = get_staff_return_core_fields(conn, active_only=True)
    core_field_keys, core_labels, core_required, core_sections = build_core_field_helpers(core_fields)
    custom_fields = get_staff_return_fields(conn, active_only=True)
    custom_answers = get_staff_return_answers(conn, existing['id']) if existing else {}
    if request.method == 'POST':
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        data = {
            'full_name': request.form.get('full_name','').strip() or user['full_name'],
            'employee_number': request.form.get('employee_number','').strip(),
            'nrc_number': request.form.get('nrc_number','').strip(),
            'gender': request.form.get('gender','').strip(),
            'date_of_birth': request.form.get('date_of_birth','').strip(),
            'phone': request.form.get('phone','').strip(),
            'email': request.form.get('email','').strip(),
            'address': request.form.get('address','').strip(),
            'position': request.form.get('position','').strip() or user['position'],
            'department': request.form.get('department','').strip(),
            'subjects_taught': request.form.get('subjects_taught','').strip(),
            'classes_taught': request.form.get('classes_taught','').strip(),
            'highest_qualification': request.form.get('highest_qualification','').strip(),
            'professional_qualification': request.form.get('professional_qualification','').strip(),
            'years_experience': request.form.get('years_experience','').strip(),
            'date_first_appointed': request.form.get('date_first_appointed','').strip(),
            'date_joined_school': request.form.get('date_joined_school','').strip(),
            'employment_status': request.form.get('employment_status','').strip(),
            'tsc_number': request.form.get('tsc_number','').strip(),
            'next_of_kin': request.form.get('next_of_kin','').strip(),
            'next_of_kin_phone': request.form.get('next_of_kin_phone','').strip(),
            'remarks': request.form.get('remarks','').strip(),
        }
        if existing:
            conn.execute("""UPDATE staff_returns SET full_name=?, employee_number=?, nrc_number=?, gender=?, date_of_birth=?, phone=?, email=?, address=?, position=?, department=?, subjects_taught=?, classes_taught=?, highest_qualification=?, professional_qualification=?, years_experience=?, date_first_appointed=?, date_joined_school=?, employment_status=?, tsc_number=?, next_of_kin=?, next_of_kin_phone=?, remarks=?, updated_at=? WHERE user_id=?""",
                         (data['full_name'], data['employee_number'], data['nrc_number'], data['gender'], data['date_of_birth'], data['phone'], data['email'], data['address'], data['position'], data['department'], data['subjects_taught'], data['classes_taught'], data['highest_qualification'], data['professional_qualification'], data['years_experience'], data['date_first_appointed'], data['date_joined_school'], data['employment_status'], data['tsc_number'], data['next_of_kin'], data['next_of_kin_phone'], data['remarks'], now, session['user_id']))
        else:
            cur = conn.execute("""INSERT INTO staff_returns(user_id, full_name, employee_number, nrc_number, gender, date_of_birth, phone, email, address, position, department, subjects_taught, classes_taught, highest_qualification, professional_qualification, years_experience, date_first_appointed, date_joined_school, employment_status, tsc_number, next_of_kin, next_of_kin_phone, remarks, submitted_at, updated_at)
                            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                         (session['user_id'], data['full_name'], data['employee_number'], data['nrc_number'], data['gender'], data['date_of_birth'], data['phone'], data['email'], data['address'], data['position'], data['department'], data['subjects_taught'], data['classes_taught'], data['highest_qualification'], data['professional_qualification'], data['years_experience'], data['date_first_appointed'], data['date_joined_school'], data['employment_status'], data['tsc_number'], data['next_of_kin'], data['next_of_kin_phone'], data['remarks'], now, now))
            existing = conn.execute('SELECT * FROM staff_returns WHERE id=?', (cur.lastrowid,)).fetchone()
        staff_return_id = existing['id'] if existing else conn.execute('SELECT id FROM staff_returns WHERE user_id=?', (session['user_id'],)).fetchone()['id']
        save_staff_return_answers(conn, staff_return_id, custom_fields, request.form)
        conn.execute('UPDATE users SET full_name=?, phone=?, email=?, address=?, qualification=?, department=? WHERE id=?',
                     (data['full_name'], data['phone'], data['email'], data['address'], data['highest_qualification'], data['department'] if data['department'] in DEPARTMENTS else user['department'], session['user_id']))
        conn.commit(); conn.close(); refresh_session_user(session['user_id'])
        flash('Staff return details saved successfully for HR records.', 'success')
        return redirect(url_for('staff_return'))
    conn.close()
    return render_template('staff_return.html', user=user, staff_return=existing, custom_fields=custom_fields, custom_answers=custom_answers, core_field_keys=core_field_keys, core_labels=core_labels, core_required=core_required)

@app.route('/staff-return-records')
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def staff_return_records():
    selected_department = request.args.get('department','')
    conn = get_db()
    if selected_department:
        rows = conn.execute("""SELECT sr.*, u.username, u.role FROM staff_returns sr JOIN users u ON sr.user_id=u.id WHERE sr.department=? ORDER BY sr.updated_at DESC, sr.full_name""", (selected_department,)).fetchall()
    else:
        rows = conn.execute("""SELECT sr.*, u.username, u.role FROM staff_returns sr JOIN users u ON sr.user_id=u.id ORDER BY sr.updated_at DESC, sr.full_name""").fetchall()
    custom_fields = get_staff_return_fields(conn, active_only=False)
    core_fields = get_staff_return_core_fields(conn, active_only=False)
    conn.close()
    return render_template('staff_return_records.html', records=rows, selected_department=selected_department, custom_fields=custom_fields, core_fields=core_fields)

@app.route('/edit-staff-return/<int:record_id>', methods=['GET','POST'])
@login_required
@roles_required('hr')
def edit_staff_return(record_id):
    """Allow HR to correct or update staff return details submitted by staff."""
    conn = get_db()
    record = conn.execute('SELECT * FROM staff_returns WHERE id=?', (record_id,)).fetchone()
    if not record:
        conn.close()
        flash('Staff return record was not found.', 'danger')
        return redirect(url_for('staff_return_records'))
    user = conn.execute('SELECT * FROM users WHERE id=?', (record['user_id'],)).fetchone()
    core_fields = get_staff_return_core_fields(conn, active_only=True)
    core_field_keys, core_labels, core_required, core_sections = build_core_field_helpers(core_fields)
    custom_fields = get_staff_return_fields(conn, active_only=True)
    custom_answers = get_staff_return_answers(conn, record_id)
    if request.method == 'POST':
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        data = {
            'full_name': request.form.get('full_name','').strip() or record['full_name'],
            'employee_number': request.form.get('employee_number','').strip(),
            'nrc_number': request.form.get('nrc_number','').strip(),
            'gender': request.form.get('gender','').strip(),
            'date_of_birth': request.form.get('date_of_birth','').strip(),
            'phone': request.form.get('phone','').strip(),
            'email': request.form.get('email','').strip(),
            'address': request.form.get('address','').strip(),
            'position': request.form.get('position','').strip(),
            'department': request.form.get('department','').strip(),
            'subjects_taught': request.form.get('subjects_taught','').strip(),
            'classes_taught': request.form.get('classes_taught','').strip(),
            'highest_qualification': request.form.get('highest_qualification','').strip(),
            'professional_qualification': request.form.get('professional_qualification','').strip(),
            'years_experience': request.form.get('years_experience','').strip(),
            'date_first_appointed': request.form.get('date_first_appointed','').strip(),
            'date_joined_school': request.form.get('date_joined_school','').strip(),
            'employment_status': request.form.get('employment_status','').strip(),
            'tsc_number': request.form.get('tsc_number','').strip(),
            'next_of_kin': request.form.get('next_of_kin','').strip(),
            'next_of_kin_phone': request.form.get('next_of_kin_phone','').strip(),
            'remarks': request.form.get('remarks','').strip(),
        }
        conn.execute("""UPDATE staff_returns SET full_name=?, employee_number=?, nrc_number=?, gender=?, date_of_birth=?, phone=?, email=?, address=?, position=?, department=?, subjects_taught=?, classes_taught=?, highest_qualification=?, professional_qualification=?, years_experience=?, date_first_appointed=?, date_joined_school=?, employment_status=?, tsc_number=?, next_of_kin=?, next_of_kin_phone=?, remarks=?, updated_at=? WHERE id=?""",
                     (data['full_name'], data['employee_number'], data['nrc_number'], data['gender'], data['date_of_birth'], data['phone'], data['email'], data['address'], data['position'], data['department'], data['subjects_taught'], data['classes_taught'], data['highest_qualification'], data['professional_qualification'], data['years_experience'], data['date_first_appointed'], data['date_joined_school'], data['employment_status'], data['tsc_number'], data['next_of_kin'], data['next_of_kin_phone'], data['remarks'], now, record_id))
        save_staff_return_answers(conn, record_id, custom_fields, request.form)
        # Keep the staff profile table aligned with the official staff return details.
        if user:
            conn.execute('UPDATE users SET full_name=?, phone=?, email=?, address=?, qualification=?, department=?, position=? WHERE id=?',
                         (data['full_name'], data['phone'], data['email'], data['address'], data['highest_qualification'], data['department'] if data['department'] in DEPARTMENTS else user['department'], data['position'] or user['position'], record['user_id']))
        conn.commit(); conn.close()
        flash('HR updated the staff return details successfully.', 'success')
        return redirect(url_for('staff_return_records'))
    conn.close()
    return render_template('staff_return.html', user=user, staff_return=record, edit_mode=True, target_name=record['full_name'], custom_fields=custom_fields, custom_answers=custom_answers, core_field_keys=core_field_keys, core_labels=core_labels, core_required=core_required)


@app.route('/manage-staff-return-details', methods=['GET','POST'])
@login_required
@roles_required('hr')
def manage_staff_return_details():
    conn = get_db()
    if request.method == 'POST':
        label = request.form.get('field_label','').strip()
        section = request.form.get('section','').strip() or 'Additional Details Required by HR'
        field_type = request.form.get('field_type','text').strip()
        options = request.form.get('options','').strip()
        is_required = request.form.get('is_required','No')
        is_active = request.form.get('is_active','Yes')
        try:
            display_order = int(request.form.get('display_order','1') or 1)
        except ValueError:
            display_order = 1
        if not label:
            flash('Please enter the staff return detail name.', 'danger')
        else:
            now = datetime.now().strftime('%Y-%m-%d %H:%M')
            conn.execute("""INSERT INTO staff_return_fields(field_label, section, field_type, options, is_required, is_active, display_order, created_by, created_at, updated_at)
                            VALUES(?,?,?,?,?,?,?,?,?,?)""",
                         (label, section, field_type, options, is_required, is_active, display_order, session.get('username'), now, now))
            conn.commit()
            flash('New staff return detail added. Teachers will see it when completing staff return.', 'success')
        return redirect(url_for('manage_staff_return_details'))
    fields = get_staff_return_fields(conn, active_only=False)
    core_fields = get_staff_return_core_fields(conn, active_only=False)
    conn.close()
    return render_template('manage_staff_return_details.html', fields=fields, core_fields=core_fields)


@app.route('/edit-core-staff-return-detail/<int:field_id>', methods=['GET','POST'])
@login_required
@roles_required('hr')
def edit_core_staff_return_detail(field_id):
    conn = get_db()
    field = conn.execute('SELECT * FROM staff_return_core_fields WHERE id=?', (field_id,)).fetchone()
    if not field:
        conn.close(); flash('Staff return detail was not found.', 'danger'); return redirect(url_for('manage_staff_return_details'))
    if request.method == 'POST':
        label = request.form.get('field_label','').strip() or field['field_label']
        section = request.form.get('section','').strip() or field['section']
        is_required = request.form.get('is_required','No')
        is_active = request.form.get('is_active','Yes')
        try:
            display_order = int(request.form.get('display_order', field['display_order']) or field['display_order'])
        except ValueError:
            display_order = field['display_order']
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        conn.execute("""UPDATE staff_return_core_fields SET field_label=?, section=?, is_required=?, is_active=?, display_order=?, updated_at=? WHERE id=?""",
                     (label, section, is_required, is_active, display_order, now, field_id))
        conn.commit(); conn.close()
        flash('Standard staff return detail updated. If inactive, teachers will no longer see it on the form.', 'success')
        return redirect(url_for('manage_staff_return_details'))
    conn.close()
    return render_template('edit_core_staff_return_detail.html', field=field)

@app.route('/toggle-core-staff-return-detail/<int:field_id>', methods=['POST'])
@login_required
@roles_required('hr')
def toggle_core_staff_return_detail(field_id):
    conn = get_db()
    field = conn.execute('SELECT * FROM staff_return_core_fields WHERE id=?', (field_id,)).fetchone()
    if field:
        new_status = 'No' if field['is_active'] == 'Yes' else 'Yes'
        conn.execute('UPDATE staff_return_core_fields SET is_active=?, updated_at=? WHERE id=?', (new_status, datetime.now().strftime('%Y-%m-%d %H:%M'), field_id))
        conn.commit()
        flash(('Removed from' if new_status == 'No' else 'Restored to') + ' the staff return form.', 'info')
    conn.close()
    return redirect(url_for('manage_staff_return_details'))

@app.route('/edit-staff-return-detail/<int:field_id>', methods=['GET','POST'])
@login_required
@roles_required('hr')
def edit_staff_return_detail(field_id):
    conn = get_db()
    field = conn.execute('SELECT * FROM staff_return_fields WHERE id=?', (field_id,)).fetchone()
    if not field:
        conn.close(); flash('Staff return detail was not found.', 'danger'); return redirect(url_for('manage_staff_return_details'))
    if request.method == 'POST':
        label = request.form.get('field_label','').strip()
        section = request.form.get('section','').strip() or 'Additional Details Required by HR'
        field_type = request.form.get('field_type','text').strip()
        options = request.form.get('options','').strip()
        is_required = request.form.get('is_required','No')
        is_active = request.form.get('is_active','Yes')
        try:
            display_order = int(request.form.get('display_order','1') or 1)
        except ValueError:
            display_order = 1
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        conn.execute("""UPDATE staff_return_fields SET field_label=?, section=?, field_type=?, options=?, is_required=?, is_active=?, display_order=?, updated_at=? WHERE id=?""",
                     (label, section, field_type, options, is_required, is_active, display_order, now, field_id))
        conn.commit(); conn.close()
        flash('Staff return detail updated successfully.', 'success')
        return redirect(url_for('manage_staff_return_details'))
    conn.close()
    return render_template('edit_staff_return_detail.html', field=field)

@app.route('/delete-staff-return-detail/<int:field_id>', methods=['POST'])
@login_required
@roles_required('hr')
def delete_staff_return_detail(field_id):
    conn = get_db()
    conn.execute('DELETE FROM staff_return_answers WHERE field_id=?', (field_id,))
    conn.execute('DELETE FROM staff_return_fields WHERE id=?', (field_id,))
    conn.commit(); conn.close()
    flash('Staff return detail deleted.', 'info')
    return redirect(url_for('manage_staff_return_details'))


@app.route('/download-staff-return-records')
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def download_staff_return_records():
    selected_department = request.args.get('department','')
    conn = get_db()
    rows, custom_fields, custom_answers = fetch_staff_return_export_data(conn, selected_department)
    conn.close()
    output = StringIO()
    writer = csv.writer(output)
    base_headers = ['Full Name','Employee Number','NRC Number','Gender','Date of Birth','Phone','Email','Address','Position','Department','Subjects Taught','Classes Taught','Highest Qualification','Professional Qualification','Years Experience','Date First Appointed','Date Joined School','Employment Status','TSC Number','Next of Kin','Next of Kin Phone','Remarks','Submitted At','Updated At']
    custom_headers = [f['field_label'] for f in custom_fields]
    writer.writerow(base_headers + custom_headers)
    for r in rows:
        base_values = [r['full_name'], r['employee_number'], r['nrc_number'], r['gender'], r['date_of_birth'], r['phone'], r['email'], r['address'], r['position'], DEPARTMENTS.get(r['department'], r['department'] or ''), r['subjects_taught'], r['classes_taught'], r['highest_qualification'], r['professional_qualification'], r['years_experience'], r['date_first_appointed'], r['date_joined_school'], r['employment_status'], r['tsc_number'], r['next_of_kin'], r['next_of_kin_phone'], r['remarks'], r['submitted_at'], r['updated_at']]
        custom_values = [custom_answers.get(r['id'], {}).get(f['id'], '') for f in custom_fields]
        writer.writerow(base_values + custom_values)
    data = BytesIO(output.getvalue().encode('utf-8'))
    data.seek(0)
    return send_file(data, mimetype='text/csv', as_attachment=True, download_name='staff_return_records.csv')

@app.route('/download-staff-return-records-excel')
@login_required
@roles_required('headteacher','deputy_headteacher','hr')
def download_staff_return_records_excel():
    selected_department = request.args.get('department','')
    conn = get_db()
    rows, custom_fields, custom_answers = fetch_staff_return_export_data(conn, selected_department)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Staff Return Records'

    title = 'KAFUBU BLOCK SECONDARY SCHOOL - STAFF RETURN RECORDS'
    if selected_department:
        title += ' - ' + DEPARTMENTS.get(selected_department, selected_department)
    end_col = get_column_letter(24 + len(custom_fields))
    ws.merge_cells(f'A1:{end_col}1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Full Name','Employee Number','NRC Number','Gender','Date of Birth','Phone','Email','Address','Position','Department','Subjects Taught','Classes Taught','Highest Qualification','Professional Qualification','Years Experience','Date First Appointed','Date Joined School','Employment Status','TSC Number','Next of Kin','Next of Kin Phone','Remarks','Submitted At','Updated At'] + [f['field_label'] for f in custom_fields]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color='1F6B3A', end_color='1F6B3A', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for r in rows:
        base_values = [
            r['full_name'], r['employee_number'], r['nrc_number'], r['gender'], r['date_of_birth'],
            r['phone'], r['email'], r['address'], r['position'],
            DEPARTMENTS.get(r['department'], r['department'] or ''),
            r['subjects_taught'], r['classes_taught'], r['highest_qualification'],
            r['professional_qualification'], r['years_experience'], r['date_first_appointed'],
            r['date_joined_school'], r['employment_status'], r['tsc_number'],
            r['next_of_kin'], r['next_of_kin_phone'], r['remarks'], r['submitted_at'], r['updated_at']
        ]
        custom_values = [custom_answers.get(r['id'], {}).get(f['id'], '') for f in custom_fields]
        ws.append(base_values + custom_values)

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for col in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 30)

    ws.freeze_panes = 'A4'
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='staff_return_records.xlsx')


@app.route('/manage-subject-teachers')
@login_required
def manage_subject_teachers():
    if session.get('role') not in STAFF_MANAGEMENT_FULL_ROLES and session.get('role') not in ROLE_TO_DEPT:
        flash('Only HODs, the Deputy Headteacher, the Headteacher and HR can manage staff profiles.', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db()
    if session.get('role') in STAFF_MANAGEMENT_FULL_ROLES:
        teachers = conn.execute("""SELECT * FROM users
                                   WHERE role='teacher' OR role LIKE 'hod_%'
                                   ORDER BY CASE WHEN role LIKE 'hod_%' THEN 1 ELSE 2 END, COALESCE(department,''), full_name""").fetchall()
        page_title = 'Manage HOD and Subject Teacher Profiles'
        dept_name = 'All Departments'
    else:
        dept = ROLE_TO_DEPT[session.get('role')]
        teachers = conn.execute("""SELECT * FROM users
                                   WHERE role='teacher' AND department=?
                                   ORDER BY full_name""", (dept,)).fetchall()
        page_title = 'Manage Subject Teachers in My Department'
        dept_name = DEPARTMENTS.get(dept, 'My Department')
    conn.close()
    return render_template('manage_subject_teachers.html', teachers=teachers, page_title=page_title, dept_name=dept_name)

@app.route('/edit-staff-profile/<int:user_id>', methods=['GET','POST'])
@login_required
def edit_staff_profile(user_id):
    conn = get_db()
    target = conn.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
    if not target:
        conn.close(); flash('Staff profile not found.', 'danger'); return redirect(url_for('manage_subject_teachers'))
    if not user_can_manage_staff(target):
        conn.close(); flash('You do not have permission to manage this staff profile.', 'danger'); return redirect(url_for('dashboard'))
    if request.method == 'POST':
        full_name = request.form['full_name'].strip()
        position = request.form.get('position','Subject Teacher').strip() or 'Subject Teacher'
        # HOD accounts remain attached to the department defined by their HOD role.
        # Subject Teacher departments can be changed by the Headteacher, Deputy Headteacher and HR.
        if target['role'].startswith('hod_'):
            department = ROLE_TO_DEPT.get(target['role'])
        elif session.get('role') in STAFF_MANAGEMENT_FULL_ROLES:
            department = request.form.get('department') or None
            if department not in DEPARTMENTS:
                department = None
        else:
            department = ROLE_TO_DEPT[session.get('role')]
        phone = request.form.get('phone','')
        email = request.form.get('email','')
        qualification = request.form.get('qualification','')
        address = request.form.get('address','')
        bio = request.form.get('bio','')
        filename = target['profile_picture']
        file = request.files.get('profile_picture')
        if file and file.filename:
            if not is_allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
                conn.close(); flash('Only JPG, PNG, GIF or WEBP profile pictures are allowed.', 'danger'); return redirect(url_for('profile'))
            if not uploaded_file_within_limit(file, MAX_IMAGE_BYTES):
                conn.close(); flash(f'Profile pictures must be {MAX_IMAGE_MB} MB or less.', 'danger'); return redirect(url_for('edit_staff_profile', user_id=user_id))
            filename = unique_safe_filename(file.filename)
            file.save(os.path.join(PROFILE_DIR, filename))
        profile_updated_at = datetime.now().strftime('%Y-%m-%d %H:%M')
        conn.execute("""UPDATE users SET full_name=?, position=?, department=?, phone=?, email=?, qualification=?, address=?, bio=?, profile_picture=?, profile_updated_at=? WHERE id=?""",
                     (full_name, position, department, phone, email, qualification, address, bio, filename, profile_updated_at, user_id))
        sync_staff_return_profile(conn, user_id, full_name, phone, email, address, position, department, qualification, updated_at=profile_updated_at)
        conn.commit(); conn.close()
        log_security_event('Staff profile reassigned or updated', username=target['username'], role=target['role'], details=f'Department: {DEPARTMENTS.get(department, department or "Not assigned")}')
        flash(f'Staff profile updated and synchronised with {DEPARTMENTS.get(department, "the department record")}.', 'success')
        return redirect(url_for('manage_subject_teachers'))
    conn.close()
    return render_template('edit_staff_profile.html', staff=target)

@app.route('/delete-staff-profile/<int:user_id>', methods=['POST'])
@login_required
def delete_staff_profile(user_id):
    conn = get_db()
    target = conn.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
    if not target:
        conn.close(); flash('Staff profile not found.', 'danger'); return redirect(url_for('manage_subject_teachers'))
    if not user_can_manage_staff(target):
        conn.close(); flash('You do not have permission to delete this staff profile.', 'danger'); return redirect(url_for('dashboard'))
    if target['id'] == session.get('user_id'):
        conn.close(); flash('You cannot delete your own account while logged in.', 'danger'); return redirect(url_for('manage_subject_teachers'))
    if session.get('role') in STAFF_MANAGEMENT_FULL_ROLES:
        if target['role'] != 'teacher' and not target['role'].startswith('hod_'):
            conn.close(); flash('Only HOD and Subject Teacher profiles can be deleted from this page.', 'danger'); return redirect(url_for('manage_subject_teachers'))
    elif target['role'] != 'teacher':
        conn.close(); flash('HODs can only delete Subject Teacher profiles under their own department.', 'danger'); return redirect(url_for('manage_subject_teachers'))
    conn.execute('DELETE FROM users WHERE id=?', (user_id,))
    conn.commit(); conn.close(); flash('Staff profile deleted successfully.', 'info'); return redirect(url_for('manage_subject_teachers'))


@app.route('/guidance-posts')
def guidance_posts():
    conn = get_db()
    posts = conn.execute('SELECT * FROM guidance_posts ORDER BY posted_at DESC').fetchall()
    conn.close()
    return render_template('guidance_posts.html', posts=posts)

@app.route('/upload-guidance-post', methods=['GET','POST'])
@login_required
@roles_required('guidance_counselling', 'headteacher', 'deputy_headteacher', 'hr')
def upload_guidance_post():
    if request.method == 'POST':
        title = request.form['title'].strip()
        category = request.form.get('category','Learner Support').strip()
        message = request.form.get('message','').strip()
        video_filename = None
        file = request.files.get('video')
        if file and file.filename:
            allowed = {'.mp4', '.webm', '.ogg', '.mov'}
            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in allowed:
                flash('Please upload a valid video file such as MP4, WEBM, OGG or MOV.', 'warning')
                return redirect(url_for('upload_guidance_post'))
            # Hosted uploads use a conservative limit to protect server memory and storage.
            file.seek(0, os.SEEK_END)
            file_size = file.tell()
            file.seek(0)
            if file_size > MAX_GUIDANCE_VIDEO_BYTES:
                flash(f'The video is too large. Guidance videos must be {MAX_GUIDANCE_VIDEO_MB} MB or less.', 'warning')
                return redirect(url_for('upload_guidance_post'))
            video_filename = unique_safe_filename(file.filename)
            file.save(os.path.join(GUIDANCE_VIDEO_DIR, video_filename))
        if not title or not message:
            flash('Title and message are required.', 'warning')
            return redirect(url_for('upload_guidance_post'))
        conn = get_db()
        conn.execute("""INSERT INTO guidance_posts(title,category,message,video_filename,posted_by,posted_at)
                        VALUES(?,?,?,?,?,?)""", (title, category, message, video_filename, session['full_name'], datetime.now().strftime('%Y-%m-%d %H:%M')))
        conn.commit(); conn.close()
        flash('Guidance and counselling post uploaded successfully.', 'success')
        return redirect(url_for('guidance_posts'))
    return render_template('upload_guidance_post.html')

@app.route('/delete-guidance-post/<int:post_id>', methods=['POST'])
@login_required
def delete_guidance_post(post_id):
    if session.get('role') not in ['guidance_counselling', 'headteacher', 'deputy_headteacher', 'hr']:
        flash('You are not allowed to delete guidance posts.', 'danger')
        return redirect(url_for('guidance_posts'))
    conn = get_db()
    post = conn.execute('SELECT * FROM guidance_posts WHERE id=?', (post_id,)).fetchone()
    if not post:
        conn.close(); flash('Post not found.', 'danger'); return redirect(url_for('guidance_posts'))
    if session.get('role') == 'guidance_counselling' and post['posted_by'] != session.get('full_name'):
        conn.close(); flash('Guidance teachers can only delete posts they uploaded.', 'danger'); return redirect(url_for('guidance_posts'))
    conn.execute('DELETE FROM guidance_posts WHERE id=?', (post_id,))
    conn.commit(); conn.close()
    flash('Guidance post deleted.', 'info')
    return redirect(url_for('guidance_posts'))

@app.route('/guidance-video/<filename>')
def guidance_video(filename):
    return send_from_directory(GUIDANCE_VIDEO_DIR, filename)

@app.route('/download/material/<filename>')
@login_required
def download_material(filename): return send_from_directory(MATERIAL_DIR, filename, as_attachment=True)
@app.route('/download/document/<filename>')
@login_required
def download_document(filename): return send_from_directory(DOCUMENT_DIR, filename, as_attachment=True)


@app.route('/school-calendar')
def school_calendar():
    conn = get_db()
    events = conn.execute("SELECT * FROM school_calendar ORDER BY event_date ASC, start_time ASC").fetchall()
    conn.close()
    return render_template('school_calendar.html', events=events)

@app.route('/manage-school-calendar', methods=['GET','POST'])
@login_required
def manage_school_calendar():
    if not can_manage_calendar():
        flash('You are not allowed to manage the school calendar.', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db()
    if request.method == 'POST':
        title = request.form.get('title','').strip()
        event_date = request.form.get('event_date','').strip()
        start_time = request.form.get('start_time','').strip()
        end_time = request.form.get('end_time','').strip()
        category = request.form.get('category','School Event').strip() or 'School Event'
        venue = request.form.get('venue','').strip()
        audience = request.form.get('audience','').strip()
        description = request.form.get('description','').strip()
        if not title or not event_date:
            conn.close(); flash('Please enter the event title and date.', 'warning'); return redirect(url_for('manage_school_calendar'))
        conn.execute("""INSERT INTO school_calendar(title,event_date,start_time,end_time,category,venue,audience,description,created_by,created_at)
                        VALUES(?,?,?,?,?,?,?,?,?,?)""",
                     (title, event_date, start_time, end_time, category, venue, audience, description, session.get('full_name'), datetime.now().strftime('%Y-%m-%d %H:%M')))
        conn.commit(); conn.close()
        flash('School calendar event added successfully.', 'success')
        return redirect(url_for('manage_school_calendar'))
    events = conn.execute('SELECT * FROM school_calendar ORDER BY event_date ASC, start_time ASC').fetchall()
    conn.close()
    return render_template('manage_school_calendar.html', events=events)

@app.route('/edit-school-calendar-event/<int:event_id>', methods=['GET','POST'])
@login_required
def edit_school_calendar_event(event_id):
    if not can_manage_calendar():
        flash('You are not allowed to edit school calendar events.', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db()
    event = conn.execute('SELECT * FROM school_calendar WHERE id=?', (event_id,)).fetchone()
    if not event:
        conn.close(); flash('Calendar event not found.', 'danger'); return redirect(url_for('manage_school_calendar'))
    if request.method == 'POST':
        conn.execute("""UPDATE school_calendar SET title=?, event_date=?, start_time=?, end_time=?, category=?, venue=?, audience=?, description=? WHERE id=?""",
                     (request.form.get('title','').strip(), request.form.get('event_date','').strip(), request.form.get('start_time','').strip(), request.form.get('end_time','').strip(), request.form.get('category','School Event').strip(), request.form.get('venue','').strip(), request.form.get('audience','').strip(), request.form.get('description','').strip(), event_id))
        conn.commit(); conn.close()
        flash('School calendar event updated successfully.', 'success')
        return redirect(url_for('manage_school_calendar'))
    conn.close()
    return render_template('edit_school_calendar_event.html', event=event)

@app.route('/delete-school-calendar-event/<int:event_id>', methods=['POST'])
@login_required
def delete_school_calendar_event(event_id):
    if not can_manage_calendar():
        flash('You are not allowed to delete school calendar events.', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db()
    conn.execute('DELETE FROM school_calendar WHERE id=?', (event_id,))
    conn.commit(); conn.close()
    flash('School calendar event deleted successfully.', 'info')
    return redirect(url_for('manage_school_calendar'))

@app.route('/manage-backgrounds', methods=['GET','POST'])
@login_required
@roles_required(*BACKGROUND_MANAGEMENT_ROLES)
def manage_backgrounds():
    if request.method == 'POST':
        action = request.form.get('action', 'upload')
        conn = get_db()
        if action == 'toggle':
            bg_id = request.form.get('bg_id')
            bg = conn.execute('SELECT * FROM website_backgrounds WHERE id=?', (bg_id,)).fetchone()
            if bg:
                new_status = 'No' if bg['is_active'] == 'Yes' else 'Yes'
                conn.execute('UPDATE website_backgrounds SET is_active=? WHERE id=?', (new_status, bg_id))
                conn.commit()
                flash('Background status updated successfully.', 'success')
            conn.close()
            return redirect(url_for('manage_backgrounds'))
        if action == 'delete':
            bg_id = request.form.get('bg_id')
            bg = conn.execute('SELECT * FROM website_backgrounds WHERE id=?', (bg_id,)).fetchone()
            if bg:
                try:
                    path = os.path.join(BACKGROUND_DIR, bg['filename'])
                    if os.path.exists(path):
                        os.remove(path)
                except OSError:
                    pass
                conn.execute('DELETE FROM website_backgrounds WHERE id=?', (bg_id,))
                conn.commit()
                flash('Background picture deleted successfully.', 'info')
            conn.close()
            return redirect(url_for('manage_backgrounds'))
        title = request.form.get('title','School Background').strip() or 'School Background'
        files = request.files.getlist('background_images')
        files = [f for f in files if f and f.filename]
        if not files:
            # Keep compatibility with older form name if a browser cached the previous page.
            single_file = request.files.get('background_image')
            files = [single_file] if single_file and single_file.filename else []
        if not files:
            conn.close(); flash('Please choose at least one background picture to upload.', 'warning'); return redirect(url_for('manage_backgrounds'))
        if len(files) > 10:
            conn.close(); flash('You can upload a maximum of 10 background pictures at once.', 'warning'); return redirect(url_for('manage_backgrounds'))

        allowed = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
        uploaded_count = 0
        skipped_files = []
        for index, file in enumerate(files, start=1):
            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in allowed:
                skipped_files.append(file.filename)
                continue
            if not uploaded_file_within_limit(file, MAX_IMAGE_BYTES):
                skipped_files.append(file.filename + f' (larger than {MAX_IMAGE_MB} MB)')
                continue
            filename = unique_safe_filename(file.filename)
            file.save(os.path.join(BACKGROUND_DIR, filename))
            picture_title = title if len(files) == 1 else f"{title} {index}"
            conn.execute('''INSERT INTO website_backgrounds(title,filename,uploaded_by,uploaded_role,is_active,uploaded_at)
                            VALUES(?,?,?,?,?,?)''',
                         (picture_title, filename, session.get('full_name'), session.get('role'), 'Yes', datetime.now().strftime('%Y-%m-%d %H:%M')))
            uploaded_count += 1

        conn.commit(); conn.close()
        if uploaded_count:
            flash(f'{uploaded_count} website background picture(s) uploaded successfully. They will rotate every 15 seconds.', 'success')
        if skipped_files:
            flash('Some files were skipped because they were not valid JPG, PNG, GIF or WEBP images: ' + ', '.join(skipped_files), 'warning')
        return redirect(url_for('manage_backgrounds'))
    conn = get_db()
    backgrounds = conn.execute('SELECT * FROM website_backgrounds ORDER BY uploaded_at DESC').fetchall()
    conn.close()
    return render_template('manage_backgrounds.html', backgrounds=backgrounds)

@app.route('/background-picture/<filename>')
def background_picture(filename):
    return send_from_directory(BACKGROUND_DIR, filename)


@app.route('/manage-passwords', methods=['GET', 'POST'])
@login_required
@roles_required('headteacher', 'hr')
def manage_passwords():
    """Headteacher and HR can reset staff/pupil portal passwords and see password change records."""
    conn = get_db()
    q = request.args.get('q', '').strip()
    role_filter = request.args.get('role', '').strip()
    sql = 'SELECT id, username, full_name, role, position, department, student_number FROM users WHERE 1=1'
    params = []
    if q:
        sql += ' AND (username LIKE ? OR full_name LIKE ? OR student_number LIKE ?)'
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if role_filter:
        sql += ' AND role=?'
        params.append(role_filter)
    sql += ' ORDER BY CASE role WHEN "headteacher" THEN 1 WHEN "deputy_headteacher" THEN 2 WHEN "hr" THEN 3 WHEN "student" THEN 9 ELSE 5 END, full_name'
    users = conn.execute(sql, params).fetchall()
    logs = conn.execute('SELECT * FROM password_audit_logs ORDER BY changed_at DESC LIMIT 100').fetchall()
    roles = conn.execute('SELECT DISTINCT role FROM users ORDER BY role').fetchall()
    conn.close()
    return render_template('manage_passwords.html', users=users, logs=logs, q=q, role_filter=role_filter, roles=roles)

@app.route('/reset-user-password/<int:user_id>', methods=['POST'])
@login_required
@roles_required('headteacher', 'hr')
def reset_user_password(user_id):
    conn = get_db()
    target = conn.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
    if not target:
        conn.close(); flash('User account not found.', 'danger'); return redirect(url_for('manage_passwords'))
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()
    note = request.form.get('note', '').strip()
    if not new_password or not confirm_password:
        conn.close(); flash('Enter the new password and confirm it.', 'warning'); return redirect(url_for('manage_passwords'))
    if new_password != confirm_password:
        conn.close(); flash('New password and confirm password do not match.', 'danger'); return redirect(url_for('manage_passwords'))
    ok, msg = password_is_strong(new_password)
    if not ok:
        conn.close(); flash(msg, 'warning'); return redirect(url_for('manage_passwords'))
    conn.execute('UPDATE users SET password=?, must_change_password=1, is_active=1 WHERE id=?', (generate_password_hash(new_password), user_id))
    log_password_action(conn, target, 'Password reset by management', note or 'Password reset by Headteacher/HR from Manage Passwords.')
    conn.commit(); conn.close()
    flash(f'Password for {target["full_name"]} has been reset successfully. The action has been recorded.', 'success')
    return redirect(url_for('manage_passwords'))

@app.route('/password-change-records')
@login_required
@roles_required('headteacher', 'hr')
def password_change_records():
    conn = get_db()
    records = conn.execute('SELECT * FROM password_audit_logs ORDER BY changed_at DESC').fetchall()
    conn.close()
    return render_template('password_change_records.html', records=records)


@app.route('/security-audit-records')
@login_required
@roles_required('headteacher', 'hr')
def security_audit_records():
    conn = get_db()
    records = conn.execute('SELECT * FROM security_audit_logs ORDER BY created_at DESC LIMIT 500').fetchall()
    conn.close()
    return render_template('security_audit_records.html', records=records)


@app.route('/backup-center', methods=['GET', 'POST'])
@login_required
@roles_required('headteacher', 'hr')
def backup_center():
    if request.method == 'POST':
        action = request.form.get('action', '')
        if action == 'create':
            filename = create_backup_archive('manual')
            log_security_event('Manual backup created', details=filename)
            flash('A complete backup was created successfully.', 'success')
            return redirect(url_for('backup_center'))
        if action == 'restore':
            filename = request.form.get('filename', '')
            confirmation = request.form.get('confirmation', '').strip()
            current_password = request.form.get('current_password', '')
            conn = get_db()
            current_user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
            conn.close()
            if confirmation != 'RESTORE':
                flash('Type RESTORE exactly to confirm.', 'warning'); return redirect(url_for('backup_center'))
            if not current_user or not check_password_hash(current_user['password'], current_password):
                flash('Your current password is incorrect.', 'danger'); return redirect(url_for('backup_center'))
            try:
                restore_backup_archive(filename)
            except (ValueError, zipfile.BadZipFile, sqlite3.DatabaseError) as exc:
                log_security_event('Backup restore failed', details=str(exc))
                flash(f'Backup restore failed: {exc}', 'danger')
                return redirect(url_for('backup_center'))
            log_security_event('Backup restored', details=filename)
            session.clear()
            flash('Backup restored successfully. Please log in again.', 'success')
            return redirect(url_for('login'))
        abort(400, description='Unknown backup action.')
    return render_template(
        'backup_center.html', backups=list_backup_archives(),
        auto_backup_enabled=AUTO_BACKUP_ENABLED, backup_retention=BACKUP_RETENTION
    )


@app.route('/backup-download/<filename>')
@login_required
@roles_required('headteacher', 'hr')
def backup_download(filename):
    safe_name = os.path.basename(filename)
    if safe_name != filename or not safe_name.endswith('.zip'):
        abort(404)
    return send_from_directory(BACKUP_DIR, safe_name, as_attachment=True)

init_db()
maybe_create_automatic_backup()


if __name__ == '__main__':
    app.run(debug=False)
